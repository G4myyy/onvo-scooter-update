import copy
import json
import os
import re
import shutil
import subprocess
import sys
import time
import threading
import unicodedata
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from PIL import Image, ImageTk
import requests
from config import config

if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
    BUNDLE_DIR = sys._MEIPASS

    # Set Tcl/Tk paths for bundled applications
    if hasattr(sys, '_MEIPASS'):
        base_path = sys._MEIPASS
    else:
        base_path = APP_DIR

    # Common locations for Tcl/Tk data in bundled applications
    tcl_paths = [
        os.path.join(base_path, '_internal', 'tcl8.6'),
        os.path.join(base_path, 'tcl8.6'),
        os.path.join(base_path, '_tcl_data'),
        r'C:\Program Files (x86)\Common Files\Microsoft\Shared\Tcl\tcl\tcl86\tcl\library',
        r'C:\Program Files\Microsoft\Windows\Tcl\tcl86\tcl\library',
    ]

    for tcl_path in tcl_paths:
        if os.path.exists(os.path.join(tcl_path, 'tcl8.6')):
            os.environ['TCL_LIBRARY'] = os.path.join(tcl_path, 'tcl8.6')
            os.environ['TK_LIBRARY'] = os.path.join(tcl_path, 'tk8.6')
            break
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
    BUNDLE_DIR = APP_DIR
EXCEL_DIR = os.path.join(APP_DIR, "excel")
DATA_FILE = os.path.join(APP_DIR, "data.json")
TEKLIF_COUNTER = os.path.join(APP_DIR, "teklif_counter.txt")
TEKLIF_DATA = os.path.join(APP_DIR, "teklifler.json")
TEKLIF_DIR = os.path.join(APP_DIR, "teklifler")

if getattr(sys, 'frozen', False):
    if not os.path.exists(DATA_FILE):
        import shutil
        bundle_data = os.path.join(BUNDLE_DIR, "data.json")
        if os.path.exists(bundle_data):
            shutil.copy2(bundle_data, DATA_FILE)
    if not os.path.isdir(EXCEL_DIR):
        import shutil
        bundle_excel = os.path.join(BUNDLE_DIR, "excel")
        if os.path.isdir(bundle_excel):
            shutil.copytree(bundle_excel, EXCEL_DIR)
COLUMNS = ("Kullanilan Urun", "Stok Adi", "Kullanilan Cihaz", "Satis Fiyati")

FIREBASE_URL = config.firebase_url
APP_VERSION = "2.67"

def _parse_version(v):
    if isinstance(v, str) and "." in v:
        parts = v.split(".")
        return (int(parts[0]), int(parts[1]))
    return (0, int(v) if str(v).isdigit() else 0)

COLUMN_HEADERS = {
    "Kullanilan Urun": "Kullanılan Ürün",
    "Stok Adi": "Stok Adı",
    "Kullanilan Cihaz": "Kullanılan Cihaz",
    "Satis Fiyati": "Satış Fiyatı",
}

TR_MAP = str.maketrans({"ı": "i", "ğ": "g", "ü": "u", "ö": "o", "ş": "s", "ç": "c"})


def normalize(text):
    if text is None:
        return ""
    text = str(text).strip().lower().translate(TR_MAP)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def record_key(record):
    return (
        normalize(record.get("Kullanilan Urun", "") or record.get("Kullanılan Ürün", "")),
        normalize(record.get("Stok Adi", "") or record.get("Stok Adı", "")),
        normalize(record.get("Kullanilan Cihaz", "") or record.get("Kullanılan Cihaz", "")),
    )


def record_get(record, field):
    return record.get(field, "") or record.get(COLUMN_HEADERS.get(field, ""), "")


def record_set(record, field, value):
    if field in record:
        record[field] = value
    elif COLUMN_HEADERS.get(field, "") in record:
        record[COLUMN_HEADERS[field]] = value
    else:
        record[field] = value


def parse_price(value):
    if value in (None, ""):
        return ""
    text = str(value).strip()
    if text.endswith("TL"):
        text = text[:-2].strip()
    if text.endswith("%"):
        text = text[:-1].strip()
    cleaned = "".join(ch for ch in text if ch.isdigit() or ch in {".", ","})
    if not cleaned:
        return ""
    if "," in cleaned and "." in cleaned:
        if cleaned.rindex(",") > cleaned.rindex("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        parts = cleaned.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    try:
        num = float(cleaned)
        if num == int(num):
            return str(int(num))
        return str(round(num, 2))
    except ValueError:
        return cleaned


def format_price(value):
    cleaned = parse_price(value)
    if not cleaned:
        return ""
    try:
        num = float(cleaned.replace(",", "."))
        return f"{num:,.2f} TL".replace(",", " ").replace(".", ",").replace(" ", ".")
    except ValueError:
        return cleaned


def find_column(headers, names):
    norm_names = [normalize(n) for n in names]
    for nn in norm_names:
        for i, h in enumerate(headers):
            if normalize(h) == nn:
                return i
    for nn in norm_names:
        for i, h in enumerate(headers):
            hn = normalize(h)
            if nn in hn and "fiyat" in nn and "fiyat" in hn:
                return i
    return None


def build_record(row, headers):
    idx_urun = find_column(headers, ["Kullanılan Ürün", "Kullanılan Ürün Adı", "Ürün Adı"])
    idx_stok = find_column(headers, ["Stok Adı", "Stok Kodu", "Ürün Kodu"])
    idx_cihaz = find_column(headers, ["Kul. Cih.", "Kullanılan Cihaz", "Cihaz"])
    idx_fiyat = 4  # E sütunu (Alış Birim Fiyat)

    def val(idx):
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v)

    return {
        "Kullanilan Urun": val(idx_urun),
        "Stok Adi": val(idx_stok),
        "Kullanilan Cihaz": val(idx_cihaz),
        "Satis Fiyati": parse_price(val(idx_fiyat)),
    }


class StokUygulamasi:
    def __init__(self):
        self.veriler = []
        self.cop_kutusu = []
        self.sepet = []
        self._sepet_pencere = None
        self._sepet_refresh = None
        self._cop_refresh = None

        self.app = tk.Tk()
        self.app.configure(bg="#F0F0F0")
        self.app.title("ONVO SCOOTER")
        w, h = 1150, 680
        sw = self.app.winfo_screenwidth()
        sh = self.app.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self._app_x = x
        self._app_y = y
        self.app.geometry(f"{w}x{h}+{x}+{y}")
        _icon = os.path.join(APP_DIR, "logo.ico")
        if not os.path.exists(_icon) and getattr(sys, 'frozen', False):
            _icon = os.path.join(BUNDLE_DIR, "logo.ico")
        if os.path.exists(_icon):
            try:
                self.app.iconbitmap(_icon)
            except Exception:
                pass
        self.app.attributes("-topmost", True)
        self.app.after(3000, lambda: self.app.attributes("-topmost", False))
        self.app.minsize(1000, 650)

        style = ttk.Style(self.app)
        style.theme_use("clam")
        style.configure("Centered.TCombobox", justify="center")

        self.arama_var = tk.StringVar()
        self.model_var = tk.StringVar(value="Tümü")
        self._secili_fiyat = ""
        self._updating_fiyat = False
        self._syncing = False
        self._log_entries = []
        self._last_push_time = 0
        self._admin_pw = "12345"
        self._logged_in = False
        self._info_overlay = None
        self._info_pages = [
            "Ürünlerinizi Excel'den içe aktarabilir,\nfiyatlara %%20 kar ekleyerek satış\nfiyatına dönüştürebilirsiniz.\n\nÜrün eklemek için ürün adı, stok adı,\ncihaz ve fiyat bilgilerini doldurup\n'Ekle / Güncelle' butonuna tıklayın.",
            "Ürünleri silmek için tablodan seçip\n'Sil' butonuna tıklayın.\n\nSilinen ürünler çöp kutusuna gider\nve 1 saat sonra kalıcı olarak silinir.\n\nÇöp kutusundan ürünleri geri alabilirsiniz.",
            "Ürünleri sepete eklemek için\ntablodan çift tıklayın veya\n'Sepete Ekle' butonunu kullanın.\n\nSepet teklif oluşturma için kullanılır.\nTeklifler Excel olarak dışa aktarılabilir.",
            "Hazır teklifleri güncelleyebilir\nve PDF olarak dışa aktarabilirsiniz.",
        ]

        self._build_ui()
        self._btn_excel.grid_remove()
        self._btn_veri.grid_remove()
        self._load_data()
        self._merge_excel_prices()
        self._refresh_models()
        self._refresh_table()
        self._log(f"Uygulama başladı ({len(self.veriler)} kayıt)")

        self.app.after(1000, self._initial_firebase_sync)
        self.app.after(2000, self._check_update)
        self.app.after(604800000, self._periodic_update_check)
        self.app.after(60000, self._periodic_cleanup)

    def _open_info_pages(self, event=None):
        if hasattr(self, '_info_overlay') and self._info_overlay:
            self._info_overlay.destroy()
            self._info_overlay = None
            return

        self._info_page_index = 0
        pw = tk.Frame(self.app, bg="white", bd=2, relief="raised")
        self._info_overlay = pw

        pw.place(x=10, y=65, width=350, height=250)
        pw.lift()

        title = tk.Label(pw, text="Kullanim Kilavuzu", font=("Arial", 14, "bold"), fg="#1F4E79", bg="white")
        title.pack(pady=(15, 10))

        close_btn = tk.Button(pw, text=" Kapat", image=self._btn_icons[8], compound="left", 
                             font=("Segoe UI Emoji", 9), bd=0, cursor="hand2", bg="white", activebackground="#f0f0f0")
        close_btn.place(x=285, y=3)
        close_btn.bind("<Button-1>", lambda e: (pw.destroy(), setattr(self, '_info_overlay', None)))

        nav_frame = tk.Frame(pw, bg="#E8F0FE")
        nav_frame.pack(side="bottom", fill="x")

        page_label = tk.Label(nav_frame, text=f"1 / {len(self._info_pages)}", 
                             font=("Arial", 10, "bold"), bg="#E8F0FE", fg="#333")
        page_label.pack(side="left", expand=True)

        def show_page(i):
            self._info_page_index = i
            self._page_text.config(text=self._info_pages[i])
            prev_btn.config(state="normal" if i > 0 else "disabled")
            next_btn.config(state="normal" if i < len(self._info_pages)-1 else "disabled")
            page_label.config(text=f"{i+1} / {len(self._info_pages)}")

        prev_btn = tk.Button(nav_frame, text="  < Geri  ", font=("Arial", 9, "bold"), 
                            command=lambda: show_page(self._info_page_index-1), state="disabled",
                            bg="#E8F0FE", fg="#1F4E79", bd=0, cursor="hand2")
        prev_btn.pack(side="left", padx=10, pady=5)

        next_btn = tk.Button(nav_frame, text="  Ileri >  ", font=("Arial", 9, "bold"), 
                            command=lambda: show_page(self._info_page_index+1),
                            bg="#E8F0FE", fg="#1F4E79", bd=0, cursor="hand2")
        next_btn.pack(side="right", padx=10, pady=5)

        self._page_text = tk.Label(pw, text=self._info_pages[0], font=("Arial", 10), bg="white", wraplength=310, justify="center", anchor="center")
        self._page_text.pack(padx=20, pady=10, fill="both", expand=True)

    def run(self):
        self.app.mainloop()

    def _build_ui(self):
        logo_path = os.path.join(APP_DIR, "logo.ico")
        if not os.path.exists(logo_path) and getattr(sys, 'frozen', False):
            logo_path = os.path.join(BUNDLE_DIR, "logo.ico")

        header = tk.Frame(self.app)
        header.pack(fill="x", padx=15, pady=(15, 10))
        header.columnconfigure(0, weight=1)
        header.columnconfigure(1, weight=0)
        header.columnconfigure(2, weight=1)

        clock_frame = tk.Frame(header, bg="#E8F0FE")
        clock_frame.grid(row=0, column=0, sticky="w")
        
        top_row = tk.Frame(clock_frame, bg="#E8F0FE")
        top_row.pack(side="top")

        self._info_btn = tk.Label(top_row, text="  ℹ  ", font=("Arial", 10, "bold"), fg="#1F4E79", bg="white", cursor="hand2", relief="solid", bd=1, padx=6, pady=2)
        self._info_btn.pack(side="left", padx=(0, 3))
        self._info_btn.bind("<Button-1>", self._open_info_pages)
        
        self._header_login = tk.Label(top_row, text=" Giriş ", font=("Arial", 9, "bold"), fg="#1F4E79", bg="#E8F0FE", cursor="hand2", padx=8, pady=2)
        self._header_login.pack(side="left", padx=(3, 0))
        self._header_login.bind("<Button-1>", lambda e: self._toggle_login())
        self._header_login.bind("<Enter>", lambda e: self._header_login.configure(bg="#D0E4FA"))
        self._header_login.bind("<Leave>", lambda e: self._header_login.configure(bg="#E8F0FE"))

        self._clock_label = tk.Label(clock_frame, text="", font=("Consolas", 12, "bold"), fg="#1F4E79", bg="#E8F0FE", padx=6, pady=3)
        self._clock_label.pack(side="top")
        self._update_clock()

        center = tk.Frame(header)
        center.grid(row=0, column=1, sticky="")

        if os.path.exists(logo_path):
            try:
                img = Image.open(logo_path)
                img = img.resize((64, 64), Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(img)
                tk.Label(center, image=self._logo_img).pack(side="left", padx=(0, 12))
            except Exception:
                pass

        tk.Label(center, text="ONVO SCOOTER FİYAT LİSTESİ", font=("Arial", 24, "bold")).pack(side="left")

        filtre = tk.Frame(self.app)
        filtre.pack(pady=(0, 6))
        tk.Label(filtre, text="Model:", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=(0, 6))
        self.model_combo = ttk.Combobox(filtre, textvariable=self.model_var, state="readonly", width=25, height=20, style="Centered.TCombobox")
        self.model_combo.grid(row=0, column=1)
        self.model_combo.bind("<<ComboboxSelected>>", self._on_model_change)

        self.arama = tk.Entry(self.app, width=60, font=("Arial", 11), textvariable=self.arama_var)
        self.arama.pack(pady=5)
        self.arama.insert(0, "Ürün ara...")
        self.arama.bind("<KeyRelease>", self._on_search)
        self.arama.bind("<FocusIn>", self._on_search_focus)
        self.arama.bind("<FocusOut>", self._on_search_blur)

        _emoji_font = ("Segoe UI Emoji", 10)

        def _make_icon(size, draw_fn):
            img = Image.new("RGBA", (size, size), (0, 0, 0, 0))
            from PIL import ImageDraw
            draw = ImageDraw.Draw(img)
            draw_fn(draw, size)
            return ImageTk.PhotoImage(img)

        def _icon_home():
            return _make_icon(18, lambda d, s: (
                d.rectangle([2, 8, 16, 16], fill="#4A90D9"),
                d.polygon([1, 9, 9, 1, 17, 9], fill="#2E6DB4"),
                d.rectangle([7, 11, 11, 16], fill="#FFFFFF"),
            ))

        def _icon_refresh():
            return _make_icon(18, lambda d, s: (
                d.arc([1, 1, 17, 17], 30, 330, fill="#27AE60", width=2),
                d.polygon([14, 0, 17, 4, 12, 4], fill="#27AE60"),
            ))

        def _icon_trash():
            return _make_icon(18, lambda d, s: (
                d.rectangle([4, 4, 14, 16], fill="#E74C3C"),
                d.rectangle([3, 2, 15, 5], fill="#C0392B"),
                d.line([7, 7, 7, 14], fill="#FFF", width=1),
                d.line([9, 7, 9, 14], fill="#FFF", width=1),
                d.line([11, 7, 11, 14], fill="#FFF", width=1),
            ))

        def _icon_excel():
            return _make_icon(18, lambda d, s: (
                d.rectangle([3, 1, 15, 17], fill="#27AE60"),
                d.rectangle([5, 4, 13, 6], fill="#FFF"),
                d.rectangle([5, 8, 13, 10], fill="#FFF"),
                d.rectangle([5, 12, 13, 14], fill="#FFF"),
            ))

        def _icon_cart():
            return _make_icon(18, lambda d, s: (
                d.polygon([2, 4, 4, 4, 6, 14, 15, 14, 16, 6, 5, 6], fill="#F39C12"),
                d.rectangle([7, 15, 14, 17], fill="#E67E22"),
                d.ellipse([8, 1, 12, 5], outline="#F39C12", width=2),
            ))

        def _icon_add():
            return _make_icon(18, lambda d, s: (
                d.rectangle([2, 2, 16, 16], fill="#3498DB"),
                d.line([9, 4, 9, 14], fill="#FFF", width=2),
                d.line([4, 9, 14, 9], fill="#FFF", width=2),
            ))

        def _icon_cart_add():
            return _make_icon(18, lambda d, s: (
                d.polygon([1, 4, 3, 4, 5, 14, 14, 14, 15, 6, 4, 6], fill="#8E44AD"),
                d.rectangle([7, 15, 13, 17], fill="#7D3C98"),
                d.line([10, 7, 10, 13], fill="#FFF", width=2),
                d.line([7, 10, 13, 10], fill="#FFF", width=2),
            ))

        def _icon_delete():
            return _make_icon(18, lambda d, s: (
                d.line([4, 4, 14, 14], fill="#E74C3C", width=3),
                d.line([14, 4, 4, 14], fill="#E74C3C", width=3),
            ))

        def _icon_close():
            return _make_icon(18, lambda d, s: (
                d.line([4, 4, 14, 14], fill="#95A5A6", width=3),
                d.line([14, 4, 4, 14], fill="#95A5A6", width=3),
            ))

        def _icon_restore():
            return _make_icon(18, lambda d, s: (
                d.arc([1, 1, 17, 17], 150, 390, fill="#2980B9", width=2),
                d.polygon([1, 5, 1, 11, 5, 8], fill="#2980B9"),
            ))

        def _icon_clean():
            return _make_icon(18, lambda d, s: (
                d.rectangle([4, 4, 14, 16], fill="#E67E22"),
                d.rectangle([3, 2, 15, 5], fill="#D35400"),
                d.rectangle([7, 10, 11, 16], fill="#FFF"),
            ))

        def _icon_offer():
            return _make_icon(18, lambda d, s: (
                d.rectangle([3, 1, 15, 17], fill="#2ECC71"),
                d.line([6, 5, 12, 5], fill="#FFF", width=1),
                d.line([6, 8, 12, 8], fill="#FFF", width=1),
                d.line([6, 11, 10, 11], fill="#FFF", width=1),
            ))

        _home_img = _icon_home()
        _refresh_img = _icon_refresh()
        _trash_img = _icon_trash()
        _excel_img = _icon_excel()
        _cart_img = _icon_cart()
        _add_img = _icon_add()
        _cart_add_img = _icon_cart_add()
        _delete_img = _icon_delete()
        _close_img = _icon_close()
        self._close_red_img = _delete_img
        _restore_img = _icon_restore()
        _clean_img = _icon_clean()
        _offer_img = _icon_offer()
        self._btn_icons = [_home_img, _refresh_img, _trash_img, _excel_img, _cart_img, _add_img, _cart_add_img, _delete_img, _close_img, _restore_img, _clean_img, _offer_img]

        btn_frame = tk.Frame(self.app)
        btn_frame.pack(pady=8)

        tk.Button(btn_frame, text=" Ana Menü", image=_home_img, compound="left", font=_emoji_font, command=self._go_home).grid(row=0, column=0, padx=5)
        tk.Button(btn_frame, text=" Yenile", image=_refresh_img, compound="left", font=_emoji_font, command=self._refresh_all).grid(row=0, column=1, padx=5)
        tk.Button(btn_frame, text=" Çöp Kutusu", image=_trash_img, compound="left", font=_emoji_font, command=self._open_trash).grid(row=0, column=2, padx=5)
        self._btn_excel = tk.Button(btn_frame, text=" Excel Seç", image=_excel_img, compound="left", font=_emoji_font, command=self._import_excel)
        self._btn_excel.grid(row=0, column=3, padx=5)
        self._btn_veri = tk.Button(btn_frame, text=" Veri Al", font=_emoji_font, command=self._reload_excel)
        self._btn_veri.grid(row=0, column=4, padx=5)
        tk.Button(btn_frame, text=" Sepet", image=_cart_img, compound="left", font=_emoji_font, command=self._open_cart).grid(row=0, column=5, padx=5)
        tk.Button(btn_frame, text=" Teklifler", image=_offer_img, compound="left", font=_emoji_font, command=self._open_teklifler).grid(row=0, column=6, padx=5)


        form = tk.LabelFrame(self.app, text="Yeni Ürün", padx=12, pady=10)
        form.pack(fill="x", padx=20, pady=(0, 10))

        labels = [("Ürün Adı", 2), ("Stok Adı", 4), ("Kullanılan Cihaz", 6), ("Satış Fiyatı", 8), ("Yüzde", 10)]
        for text, col in labels:
            tk.Label(form, text=text, font=("Arial", 10, "bold")).grid(row=0, column=col, sticky="w", padx=(0, 5), pady=3)

        self.urun_entry = tk.Entry(form, width=20)
        self.stok_entry = tk.Entry(form, width=20)
        self.cihaz_entry = tk.Entry(form, width=20)
        self.fiyat_entry = tk.Entry(form, width=20)
        self.yuzde_var = tk.StringVar()
        self.yuzde_combo = ttk.Combobox(form, textvariable=self.yuzde_var, state="readonly", width=8, values=["", "%10", "%20"])

        self.urun_entry.grid(row=1, column=2, padx=5, pady=3)
        self.stok_entry.grid(row=1, column=4, padx=5, pady=3)
        self.cihaz_entry.grid(row=1, column=6, padx=5, pady=3)
        self.fiyat_entry.grid(row=1, column=8, padx=5, pady=3)
        self.yuzde_combo.grid(row=1, column=10, padx=5, pady=3)
        self.yuzde_combo.bind("<<ComboboxSelected>>", self._on_yuzde_change)
        self.fiyat_entry.bind("<KeyRelease>", self._on_fiyat_manual)
        self.fiyat_entry.bind("<Return>", lambda e: self._add_or_update())

        _small_btn = {"font": ("Arial", 7), "width": 1, "height": 1, "bd": 0, "padx": 0, "pady": 0, "cursor": "hand2", "fg": "#999"}
        def _clear_all():
            self.urun_entry.delete(0, tk.END)
            self.stok_entry.delete(0, tk.END)
            self.cihaz_entry.delete(0, tk.END)
            self.fiyat_entry.delete(0, tk.END)
            self.yuzde_var.set("")
            self._secili_fiyat = ""
        tk.Button(form, text="×", command=lambda: self.urun_entry.delete(0, tk.END), **_small_btn).grid(row=1, column=3, sticky="w")
        tk.Button(form, text="×", command=lambda: self.stok_entry.delete(0, tk.END), **_small_btn).grid(row=1, column=5, sticky="w")
        tk.Button(form, text="×", command=lambda: self.cihaz_entry.delete(0, tk.END), **_small_btn).grid(row=1, column=7, sticky="w")
        tk.Button(form, text="×", command=lambda: self.fiyat_entry.delete(0, tk.END), **_small_btn).grid(row=1, column=9, sticky="w")
        tk.Button(form, text="×", command=_clear_all, **_small_btn).grid(row=0, column=0, sticky="w", padx=(0, 2))

        islem = tk.Frame(self.app)
        islem.pack(pady=(0, 10))
        tk.Button(islem, text=" Ekle / Güncelle", image=_add_img, compound="left", font=_emoji_font, command=self._add_or_update).grid(row=0, column=0, padx=5)
        tk.Button(islem, text=" Sepete Ekle", image=_cart_add_img, compound="left", font=_emoji_font, command=self._add_to_cart).grid(row=0, column=1, padx=5)
        tk.Button(islem, text=" Sil", image=_delete_img, compound="left", font=_emoji_font, command=self._delete).grid(row=0, column=2, padx=5)
        tk.Button(islem, text=" Kapat", image=_close_img, compound="left", font=_emoji_font, command=self.app.destroy).grid(row=0, column=3, padx=5)

        tablo_frame = tk.Frame(self.app)
        tablo_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        display_cols = list(COLUMN_HEADERS.keys())
        self.tablo = ttk.Treeview(tablo_frame, columns=display_cols, show="headings", height=13, selectmode="extended")
        for col in display_cols:
            self.tablo.heading(col, text=COLUMN_HEADERS[col])
            self.tablo.column(col, width=170 if col == "Kullanilan Urun" else 140, anchor="center")

        self.tablo.grid(row=0, column=0, sticky="nsew")
        self.tablo.tag_configure("flash", background="#3498DB", foreground="white")
        scrollbar = ttk.Scrollbar(tablo_frame, orient="vertical", command=self.tablo.yview)
        self.tablo.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns")
        tablo_frame.rowconfigure(0, weight=1)
        tablo_frame.columnconfigure(0, weight=1)

        self._last_selection = None
        self.tablo.bind("<<TreeviewSelect>>", self._on_select)
        self.tablo.bind("<Double-1>", self._on_double_click)

        self._auto_scrolling = False
        self._scroll_start_y = 0
        self._scroll_start_pos = 0

        def _start_auto_scroll(event):
            self._auto_scrolling = True
            self._scroll_start_y = event.y_root
            self._scroll_start_pos = self.tablo.yview()[0]
            self.tablo.config(cursor="sb_v_double_arrow")

        def _do_auto_scroll(event):
            if not self._auto_scrolling:
                return
            dy = event.y_root - self._scroll_start_y
            total_lines = max(1, len(self.tablo.get_children()))
            new_pos = self._scroll_start_pos + (dy / 800)
            new_pos = max(0, min(1, new_pos))
            self.tablo.yview_moveto(new_pos)

        def _stop_auto_scroll(event):
            self._auto_scrolling = False
            self.tablo.config(cursor="")

        self.tablo.bind("<Button-2>", _start_auto_scroll)
        self.tablo.bind("<B2-Motion>", _do_auto_scroll)
        self.tablo.bind("<ButtonRelease-2>", _stop_auto_scroll)

        self.tablo_alt = tk.Label(self.app, text="", font=("Arial", 10, "bold"), fg="#1F4E79")
        self.tablo_alt.pack(pady=(2, 0))

        alt = tk.Frame(self.app)
        alt.pack(fill="x", padx=20, pady=(0, 5))
        tk.Label(alt, text=f"v{APP_VERSION}", font=("Arial", 9), fg="#999").pack(side="right")
        ttk.Button(alt, text="📋 İşlem Günlüğü", command=self._open_log).pack(side="right")
        self.durum = tk.Label(alt, text="Yükleniyor...", font=("Arial", 10))
        self.durum.pack(side="left")

    def _open_log(self):
        if hasattr(self, '_log_pencere') and self._log_pencere and self._log_pencere.winfo_exists():
            self._log_pencere.lift()
            return

        self._log_pencere = tk.Toplevel(self.app)
        self._log_pencere.title("İşlem Günlüğü")
        self._log_pencere.geometry("600x450")

        self.log_text = tk.Text(self._log_pencere, font=("Consolas", 10), state="disabled", bg="#f5f5f5")
        self.log_text.pack(fill="both", expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        if hasattr(self, '_log_buffer'):
            self.log_text.config(state="normal")
            self.log_text.insert("end", self._log_buffer)
            self.log_text.see("end")
            self.log_text.config(state="disabled")

        self.log_text.bind("<Double-Button-1>", self._on_log_double_click)

        bilgi = tk.Label(self._log_pencere, text="Satıra çift tıklayarak o ana dönebilirsiniz",
                         font=("Arial", 9, "italic"), fg="#666")
        bilgi.pack(pady=(0, 5))

    def _on_log_double_click(self, event=None):
        idx = self.log_text.index(f"@{event.x},{event.y}")
        line = int(idx.split(".")[0]) - 1
        if line < 0 or line >= len(self._log_entries):
            return
        entry = self._log_entries[line]
        msg = entry["msg"]
        if not messagebox.askyesno("Geri Dön", f"Bu noktaya dönmek istiyor musunuz?\n\n{msg}"):
            return
        self.veriler = copy.deepcopy(entry["veriler"])
        self.cop_kutusu = copy.deepcopy(entry["cop"])
        self.sepet = copy.deepcopy(entry["sepet"])
        self._save()
        self._refresh_models()
        self._refresh_table()
        self._log(f"Geri dönüldü: {msg}", snapshot=False)

    def _log(self, msg, snapshot=True):
        zaman = time.strftime("%H:%M:%S")
        satir = f"[{zaman}] {msg}\n"
        if not hasattr(self, '_log_buffer'):
            self._log_buffer = ""
        self._log_buffer += satir
        lines = self._log_buffer.count("\n")
        if lines > 25:
            self._log_buffer = "".join(self._log_buffer.split("\n")[-20:])
            self._log_entries = self._log_entries[-20:]
        if snapshot:
            self._log_entries.append({
                "msg": satir.strip(),
                "veriler": copy.deepcopy(self.veriler),
                "cop": copy.deepcopy(self.cop_kutusu),
                "sepet": copy.deepcopy(self.sepet),
            })
        if hasattr(self, 'log_text') and self.log_text.winfo_exists():
            self.log_text.config(state="normal")
            self.log_text.insert("end", satir)
            self.log_text.see("end")
            self.log_text.config(state="disabled")

    def _notify(self, title, msg):
        ps = f'''
[Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, ContentType = WindowsRuntime] | Out-Null
[Windows.Data.Xml.Dom.XmlDocument, Windows.Data.Xml.Dom.XmlDocument, ContentType = WindowsRuntime] | Out-Null
$template = @"
<toast>
    <visual>
        <binding template="ToastGeneric">
            <text>{title}</text>
            <text>{msg}</text>
        </binding>
    </visual>
</toast>
"@
$xml = New-Object Windows.Data.Xml.Dom.XmlDocument
$xml.LoadXml($template)
$toast = [Windows.UI.Notifications.ToastNotification]::new($xml)
[Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier("ONVO Stok").Show($toast)
'''
        def _run():
            try:
                si = subprocess.STARTUPINFO()
                si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                subprocess.run(["powershell", "-Command", ps], capture_output=True, timeout=5,
                               startupinfo=si, creationflags=0x08000000)
            except Exception:
                pass
        threading.Thread(target=_run, daemon=True).start()

    def _on_search_focus(self, event=None):
        txt = self.arama_var.get().strip()
        if txt == "Ürün ara...":
            self.arama_var.set("")

    def _on_search_blur(self, event=None):
        if not self.arama_var.get().strip():
            self.arama_var.set("Ürün ara...")

    def _on_search(self, event=None):
        txt = self.arama_var.get().strip()
        if txt == "Ürün ara..." or not txt:
            return
        arama = self.arama_var.get().strip()
        if arama:
            self._log(f"Arama: {arama}")
        self._refresh_table()

    def _on_model_change(self, event=None):
        secim = self.model_combo.get()
        self._log(f"Model filtresi: {secim}")
        self._refresh_table()

    def _on_fiyat_manual(self, event=None):
        if self._updating_fiyat:
            return
        val = self.fiyat_entry.get().strip()
        if val:
            self._secili_fiyat = val

    def _on_yuzde_change(self, event=None):
        yuzde = self.yuzde_var.get()
        self._log(f"Yüzde seçildi: {yuzde}")
        if not self._secili_fiyat:
            return
        try:
            fiyat = float(self._secili_fiyat.replace(",", "."))
        except ValueError:
            return
        if yuzde == "%10":
            yeni = round(fiyat * 1.10, 2)
        elif yuzde == "%20":
            yeni = round(fiyat * 1.20, 2)
        else:
            yeni = fiyat
        self.fiyat_entry.delete(0, tk.END)
        self._updating_fiyat = True
        self.fiyat_entry.insert(0, f"{yeni:.2f}")
        self._updating_fiyat = False

    def _on_select(self, event=None):
        sel = self.tablo.selection()
        if not sel:
            self.tablo_alt.config(text=f"Toplam Kayıt: {len(self.veriler)}")
            self._update_cart_count()
            return
        if len(sel) > 1:
            self.tablo_alt.config(text=f"Toplam Kayıt: {len(self.veriler)}  |  Seçili: {len(sel)} ürün")
            self._update_cart_count()
            return
        self._last_selection = sel[0]
        vals = self.tablo.item(sel[0], "values")
        if not vals:
            return
        self._log(f"Seçildi: {vals[0]} - {vals[1]}")
        self.urun_entry.delete(0, tk.END)
        self.stok_entry.delete(0, tk.END)
        self.cihaz_entry.delete(0, tk.END)
        self.fiyat_entry.delete(0, tk.END)
        self.urun_entry.insert(0, vals[0])
        self.stok_entry.insert(0, vals[1])
        self.cihaz_entry.insert(0, vals[2])
        self.yuzde_var.set("")
        self._secili_fiyat = ""
        rec = self._find_record(vals[0], vals[1], vals[2])
        if rec:
            self._secili_fiyat = record_get(rec, "Satis Fiyati") or ""
            if self._secili_fiyat:
                self.fiyat_entry.insert(0, self._secili_fiyat)

    def _update_cart_count(self):
        sel_count = len(self.tablo.selection())
        cart_count = len(self.sepet)
        parts = [f"Toplam: {len(self.veriler)}"]
        if sel_count > 0:
            parts.append(f"Seçili: {sel_count}")
        if cart_count > 0:
            parts.append(f"Sepet: {cart_count}")
        self.tablo_alt.config(text="  |  ".join(parts))

    def _on_double_click(self, event=None):
        sel = self.tablo.selection()
        if not sel:
            return
        vals = self.tablo.item(sel[0], "values")
        if not vals:
            return
        urun = vals[0]
        stok = vals[1]
        cihaz = vals[2]
        rec = self._find_record(urun, stok, cihaz)
        if not rec:
            return
        fiyat_str = record_get(rec, "Satis Fiyati") or ""
        if not fiyat_str:
            return
        temiz_fiyat = parse_price(fiyat_str)
        if not temiz_fiyat or temiz_fiyat == "0":
            return
        birim_fiyat = float(temiz_fiyat.replace(",", "."))
        toplam = round(birim_fiyat * 1, 2)
        self.sepet.append({
            "Kullanilan Urun": urun,
            "Stok Adi": stok,
            "Kullanilan Cihaz": cihaz,
            "Birim Fiyat": birim_fiyat,
            "Miktar": 1,
            "Toplam": toplam,
        })
        self._log(f"Sepete eklendi: {stok} x1")
        self._notify("Sepete Eklendi", f"{stok} x1 - {toplam:,.2f} TL")
        if self._sepet_refresh:
            self._sepet_refresh()

        iid = sel[0]
        self.tablo.item(iid, tags=("flash",))
        self.app.after(200, lambda: self.tablo.item(iid, tags=()))

    def _save(self):
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump({"veriler": self.veriler, "cop_kutusu": self.cop_kutusu}, f, ensure_ascii=False, indent=2)
        self._last_push_time = time.time()
        if hasattr(self, '_push_after_id') and self._push_after_id:
            self.app.after_cancel(self._push_after_id)
        self._push_after_id = self.app.after(2000, self._firebase_push)

    def _firebase_push(self):
        if self._syncing:
            return
        self._syncing = True
        try:
            local_sepet = copy.deepcopy(self.sepet)
            requests.put(f"{FIREBASE_URL}/veriler.json", json=self.veriler, timeout=10)
            requests.put(f"{FIREBASE_URL}/cop_kutusu.json", json=self.cop_kutusu, timeout=10)
            requests.put(f"{FIREBASE_URL}/sepet.json", json=local_sepet, timeout=10)
            teklifler = []
            if os.path.exists(TEKLIF_DATA):
                try:
                    with open(TEKLIF_DATA, "r", encoding="utf-8") as f:
                        teklifler = json.load(f)
                except Exception:
                    pass
            requests.put(f"{FIREBASE_URL}/teklifler.json", json=teklifler, timeout=10)
            self.app.after(0, lambda: (self.durum.config(text=f"Bulut senkronize | Toplam: {len(self.veriler)}"), self._log("Buluta gönderildi")))
        except Exception:
            self.app.after(0, lambda: self._log("Bulut hatası!"))
        finally:
            self._syncing = False

    def _firebase_pull_local(self):
        if self._syncing or time.time() - self._last_push_time < 10:
            return False
        try:
            r1 = requests.get(f"{FIREBASE_URL}/veriler.json", timeout=10)
            r2 = requests.get(f"{FIREBASE_URL}/cop_kutusu.json", timeout=10)
            r3 = requests.get(f"{FIREBASE_URL}/sepet.json", timeout=10)
            r4 = requests.get(f"{FIREBASE_URL}/teklifler.json", timeout=10)
            if r1.status_code == 200 and r2.status_code == 200:
                remote_veriler = r1.json()
                remote_cop = r2.json()
                if not isinstance(remote_veriler, list):
                    remote_veriler = []
                if not isinstance(remote_cop, list):
                    remote_cop = []
                remote_keys = {self._build_key(r) for r in remote_veriler}
                local_lookup = {self._build_key(r): r for r in self.veriler}
                self.veriler = [r for r in self.veriler if self._build_key(r) in remote_keys]
                for r in remote_veriler:
                    k = self._build_key(r)
                    if k in local_lookup:
                        remote_price = parse_price(r.get("Satis Fiyati", ""))
                        if remote_price and remote_price != "0":
                            local_lookup[k]["Satis Fiyati"] = remote_price
                    else:
                        self.veriler.append(r)
                        local_lookup[k] = r
                remote_cop_keys = {self._build_key(r) for r in remote_cop}
                self.cop_kutusu = [r for r in self.cop_kutusu if self._build_key(r) in remote_cop_keys]
                local_cop_keys = {self._build_key(r) for r in self.cop_kutusu}
                for r in remote_cop:
                    k = self._build_key(r)
                    if k not in local_cop_keys:
                        self.cop_kutusu.append(r)
                        local_cop_keys.add(k)
                remote_sepet = r3.json() if r3.status_code == 200 else []
                if r4.status_code == 200:
                    remote_teklifler = r4.json()
                    if isinstance(remote_teklifler, list) and remote_teklifler:
                        local_teklifler = []
                        if os.path.exists(TEKLIF_DATA):
                            try:
                                with open(TEKLIF_DATA, "r", encoding="utf-8") as f:
                                    local_teklifler = json.load(f)
                            except Exception:
                                pass
                        local_no_set = {t.get("no") for t in local_teklifler}
                        for t in remote_teklifler:
                            if t.get("no") not in local_no_set:
                                local_teklifler.append(t)
                        with open(TEKLIF_DATA, "w", encoding="utf-8") as f:
                            json.dump(local_teklifler, f, ensure_ascii=False, indent=2)
                with open(DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump({"veriler": self.veriler, "cop_kutusu": self.cop_kutusu, "sepet": self.sepet}, f, ensure_ascii=False, indent=2)
                return True
        except Exception:
            pass
        return False

    def _initial_firebase_sync(self):
        def sync():
            if self._firebase_pull_local():
                self.app.after(0, self._on_firebase_sync)
            else:
                threading.Thread(target=self._firebase_push, daemon=True).start()
        threading.Thread(target=sync, daemon=True).start()
        self.app.after(5000, self._periodic_firebase_sync)

    def _periodic_firebase_sync(self):
        def sync():
            if self._firebase_pull_local():
                self.app.after(0, self._on_firebase_sync)
        threading.Thread(target=sync, daemon=True).start()
        self.app.after(5000, self._periodic_firebase_sync)

    def _on_firebase_sync(self):
        self._refresh_models()
        self._refresh_table()
        self._log(f"Buluttan güncellendi ({len(self.veriler)} kayıt)")
        self.durum.config(text=f"Bulut senkronize | Toplam: {len(self.veriler)}")

    def _check_update(self):
        def check():
            try:
                self._log("Versiyon kontrolü yapılıyor...")
                r = requests.get(f"{FIREBASE_URL}/app_version.json", timeout=10)
                self._log(f"Firebase response: {r.status_code}")
                if r.status_code != 200:
                    self._log(f"Firebase hata: {r.status_code}")
                    return
                data = r.json()
                self._log(f"Firebase data: {data}")
                if isinstance(data, dict):
                    remote_ver = data.get("version", 0)
                    download_url = data.get("url", "")
                elif isinstance(data, (int, float)):
                    remote_ver = int(data)
                    download_url = ""
                else:
                    self._log(f"Beklenmeyen data formatı: {type(data)}")
                    return
                self._log(f"Remote: {remote_ver}, Local: {APP_VERSION}, URL: {download_url}")
                if _parse_version(remote_ver) > _parse_version(APP_VERSION) and download_url:
                    self.app.after(0, lambda: self._do_update(download_url, remote_ver))
                else:
                    self._log("Güncelleme yok veya URL boş")
            except Exception as e:
                self._log(f"Versiyon kontrolü hatası: {e}")
        threading.Thread(target=check, daemon=True).start()

    def _do_update(self, url, new_ver):
        if not messagebox.askyesno("Güncelleme", f"Yeni versiyon mevcut (v{new_ver}).\nİndirilsin mi?"):
            self._log("Güncelleme iptal edildi (Hayır basıldı)")
            return
        self._log(f"Güncelleme indiriliyor: v{new_ver}")

        def download():
            try:
                self._log("Download thread started")
                r = requests.get(url, timeout=120, stream=True)
                self._log(f"Download response: {r.status_code}")
                if r.status_code != 200:
                    self._log(f"Download failed: HTTP {r.status_code}")
                    self.app.after(0, lambda: messagebox.showerror("Hata", "İndirme başarısız."))
                    return
                total = int(r.headers.get("Content-Length", 0))
                downloaded = 0
                # Download to temp directory (not app directory!)
                temp_dir = os.path.join(os.environ.get("TEMP", "."), "OnvoScooterUpdate")
                os.makedirs(temp_dir, exist_ok=True)
                tmp_exe = os.path.join(temp_dir, "OnvoScooter.exe")
                with open(tmp_exe, "wb") as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total > 0:
                            pct = round(downloaded * 100 / total)
                        else:
                            pct = 0
                        self.app.after(0, lambda p=pct: self.durum.config(text=f"Güncelleme indiriliyor... %{p}"))

                self._log(f"Download complete, file size: {os.path.getsize(tmp_exe)}")
                self.app.after(0, lambda: self.durum.config(text="Güncelleme hazırlanıyor..."))

                # Batch file for atomic replace (from temp to app dir)
                bat_path = os.path.join(APP_DIR, "updater.bat")
                new_exe = os.path.join(APP_DIR, "OnvoScooter.exe")
                bak_exe = new_exe + ".bak"

                bat_content = f'''@echo off
timeout /t 2 /nobreak >nul
move /y "{new_exe}" "{new_exe}.bak" >nul 2>&1
move /y "{tmp_exe}" "{new_exe}" >nul 2>&1
if errorlevel 1 (
    move /y "{new_exe}.bak" "{new_exe}" >nul 2>&1
) else (
    start "" "{new_exe}"
)
del "{new_exe}.bak" 2>nul
del "%~f0" 2>nul
'''

                with open(bat_path, "w", encoding="utf-8") as f:
                    f.write(bat_content)

                self._log(f"Updater batch written: {bat_path}")

                # Batch'i ayrı süreçte başlat
                proc = subprocess.Popen(["cmd", "/c", bat_path], shell=True)
                self._log(f"Updater batch launched with PID {proc.pid}")
                self.app.after(500, self.app.destroy)

            except Exception as e:
                self._log(f"Download/Updater error: {e}")
                self.app.after(0, lambda: messagebox.showerror("Hata", f"Güncelleme başarısız:\n{e}"))
        self._log("Starting download thread")
        threading.Thread(target=download, daemon=True).start()

    def _periodic_update_check(self):
        self._check_update()
        self.app.after(604800000, self._periodic_update_check)

    @staticmethod
    def _markup(price):
        if not price:
            return price
        try:
            return str(round(float(price.replace(",", ".")) * 1.2, 2))
        except ValueError:
            return price

    def _load_data(self):
        if not os.path.exists(DATA_FILE):
            return
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                self.veriler = data.get("veriler", [])
                self.cop_kutusu = data.get("cop_kutusu", [])
            elif isinstance(data, list):
                self.veriler = data
        except Exception:
            self.veriler = []
            self.cop_kutusu = []

    def _merge_excel_prices(self):
        if not os.path.isdir(EXCEL_DIR):
            return
        lookup = {self._build_key(r): r for r in self.veriler}
        for dosya in sorted(os.listdir(EXCEL_DIR)):
            if not dosya.lower().endswith(".xlsx") or dosya.startswith("~$"):
                continue
            try:
                wb = load_workbook(os.path.join(EXCEL_DIR, dosya), data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = ["" if c is None else str(c).strip() for c in rows[0]]
                for row in rows[1:]:
                    if not row or all(c is None or str(c).strip() == "" for c in row):
                        continue
                    rec = build_record(row, headers)
                    k = self._build_key(rec)
                    if k in lookup:
                        existing = lookup[k]
                        old_price = parse_price(record_get(existing, "Satis Fiyati"))
                        if not old_price or old_price == "0":
                            excel_price = rec.get("Satis Fiyati", "")
                            if excel_price:
                                record_set(existing, "Satis Fiyati", self._markup(excel_price))
            except Exception:
                continue

    def _build_key(self, record):
        urun = normalize(record_get(record, "Kullanilan Urun"))
        stok = normalize(record_get(record, "Stok Adi"))
        cihaz = normalize(record_get(record, "Kullanilan Cihaz"))
        watt = ""
        raw_urun = record_get(record, "Kullanilan Urun") or ""
        m = re.search(r'(\d+)\s*[Ww]', raw_urun)
        if m:
            watt = m.group(1)
        return (urun, stok, cihaz, watt)

    def _find_record(self, urun, stok, cihaz):
        un = normalize(urun)
        sn = normalize(stok)
        cn = normalize(cihaz)
        watt = ""
        m = re.search(r'(\d+)\s*[Ww]', str(urun))
        if m:
            watt = m.group(1)
        target = (un, sn, cn, watt)
        for r in self.veriler:
            if self._build_key(r) == target:
                return r
        return None

    def _select_by_values(self, urun, stok, cihaz):
        for item in self.tablo.get_children():
            vals = self.tablo.item(item, "values")
            if vals and vals[0] == urun and vals[1] == stok and vals[2] == cihaz:
                self.tablo.selection_set(item)
                self.tablo.see(item)
                return

    def _refresh_table(self, clear_sel=False):
        saved_sel = []
        if not clear_sel:
            for item_id in self.tablo.selection():
                vals = self.tablo.item(item_id, "values")
                if vals:
                    saved_sel.append((vals[0], vals[1], vals[2]))

        for item in self.tablo.get_children():
            self.tablo.delete(item)

        arama = self.arama_var.get().strip().lower()
        if arama in ("", "ürün ara..."):
            arama = ""

        model = self.model_var.get()
        filtered = []
        for v in self.veriler:
            if arama:
                text = " ".join(normalize(val) for val in v.values())
                flat_text = text.replace(" ", "")
                arama_kelimeler = normalize(arama).split()
                if not all(k in flat_text for k in arama_kelimeler):
                    continue
            if model and model != "Tümü":
                match_w = re.search(r'(\d+)\s*[Ww]$', model)
                if match_w:
                    watt = match_w.group(1)
                    ci_name = model[:match_w.start()].strip()
                    if normalize(record_get(v, "Kullanilan Cihaz")) != normalize(ci_name):
                        continue
                    urun = record_get(v, "Kullanilan Urun") or ""
                    if not re.search(rf'{watt}\s*[Ww]', urun):
                        continue
                else:
                    if normalize(record_get(v, "Kullanilan Cihaz")) != normalize(model):
                        continue
            filtered.append(v)

        grouped = {}
        for v in filtered:
            key = normalize(record_get(v, "Kullanilan Urun"))
            grouped.setdefault(key, []).append(v)

        ordered = []
        for key in sorted(grouped):
            ordered.extend(grouped[key])

        for idx, v in enumerate(ordered):
            fiyat = record_get(v, "Satis Fiyati")
            iid = self.tablo.insert("", "end", iid=str(idx), values=(
                record_get(v, "Kullanilan Urun"),
                record_get(v, "Stok Adi"),
                record_get(v, "Kullanilan Cihaz"),
                format_price(fiyat) if fiyat else "",
            ))
            if not clear_sel:
                vals = (record_get(v, "Kullanilan Urun"), record_get(v, "Stok Adi"), record_get(v, "Kullanilan Cihaz"))
                if vals in saved_sel:
                    self.tablo.selection_add(iid)

        self.durum.config(text=f"Gösterilen kayıt: {len(ordered)} / Toplam: {len(self.veriler)}")
        self.tablo_alt.config(text=f"Toplam Kayıt: {len(self.veriler)}")

    def _refresh_models(self):
        def model_sort_key(m):
            nums = ''.join(c for c in m if c.isdigit())
            return int(nums) if nums else 0

        model_wattages = {}

        model_wattages = {}
        for r in self.veriler:
            ci = record_get(r, "Kullanilan Cihaz")
            if not ci:
                continue
            urun = record_get(r, "Kullanilan Urun") or ""
            match = re.search(r'(\d+)\s*[Ww]', urun)
            watt = match.group(1) if match else ""
            key = (ci, watt)
            if key not in model_wattages:
                model_wattages[key] = ci + ("  " + watt + "W" if watt else "")

        modeller = sorted(model_wattages.values(), key=lambda x: x.lower())
        self.model_combo["values"] = ["Tümü"] + modeller
        if self.model_var.get() not in ["Tümü"] + modeller:
            self.model_var.set("Tümü")

    def _refresh_all(self):
        self._cleanup_trash()
        self._save()
        self.arama_var.set("")
        self.model_var.set("Tümü")
        self.urun_entry.delete(0, tk.END)
        self.stok_entry.delete(0, tk.END)
        self.cihaz_entry.delete(0, tk.END)
        self.fiyat_entry.delete(0, tk.END)
        self.yuzde_var.set("")
        self._secili_fiyat = ""
        self._refresh_models()
        self._refresh_table(clear_sel=True)
        self._log("Yenile - veriler yenilendi")
        self.app.focus_set()

    def _update_clock(self):
        tarih = time.strftime("%d.%m.%Y")
        saat = time.strftime("%H:%M:%S")
        self._clock_label.config(text=f"  {tarih}  |  {saat}  ")
        self.app.after(1000, self._update_clock)

    def _go_home(self):
        self.model_var.set("Tümü")
        self.arama_var.set("Ürün ara...")
        self._last_selection = None
        self._refresh_table(clear_sel=True)
        self.urun_entry.delete(0, tk.END)
        self.stok_entry.delete(0, tk.END)
        self.cihaz_entry.delete(0, tk.END)
        self.fiyat_entry.delete(0, tk.END)
        self.yuzde_var.set("")
        self._secili_fiyat = ""
        self.app.focus_set()
        self._log("Ana menüye dönüldü")

    def _add_or_update(self):
        urun = self.urun_entry.get().strip()
        stok = self.stok_entry.get().strip()
        cihaz = self.cihaz_entry.get().strip()
        fiyat = self.fiyat_entry.get().strip()

        if not urun or not stok:
            messagebox.showwarning("Uyarı", "Ürün adı ve stok adı zorunludur.")
            return

        temiz_fiyat = parse_price(fiyat)
        if not temiz_fiyat:
            messagebox.showwarning("Uyarı", "Lütfen geçerli bir fiyat girin.")
            return

        existing = self._find_record(urun, stok, cihaz)
        if existing:
            eski_fiyat = parse_price(record_get(existing, "Satis Fiyati"))
            if temiz_fiyat and parse_price(temiz_fiyat) != eski_fiyat:
                record_set(existing, "Satis Fiyati", temiz_fiyat)
                self._log(f"Fiyat güncellendi: {stok} = {format_price(temiz_fiyat)}")
            else:
                self._log(f"Güncelleme yok: {stok}")
            self._save()
            self._refresh_table()
            self._select_by_values(urun, stok, cihaz)
            messagebox.showinfo("Başarılı", "Ürün güncellendi.")
            return

        self.veriler.append({
            "Kullanilan Urun": urun,
            "Stok Adi": stok,
            "Kullanilan Cihaz": cihaz,
            "Satis Fiyati": temiz_fiyat,
        })
        self._log(f"Yeni ürün eklendi: {stok}")
        self._save()
        self._refresh_models()
        self._refresh_table()
        self._select_by_values(urun, stok, cihaz)
        messagebox.showinfo("Başarılı", "Yeni ürün eklendi.")

    def _delete(self):
        urun = self.urun_entry.get().strip()
        stok = self.stok_entry.get().strip()
        cihaz = self.cihaz_entry.get().strip()

        if not urun or not stok:
            messagebox.showwarning("Uyarı", "Silmek için tablodan bir ürün seçin.")
            return

        idx = -1
        watt = ""
        m = re.search(r'(\d+)\s*[Ww]', urun)
        if m:
            watt = m.group(1)
        target = (normalize(urun), normalize(stok), normalize(cihaz), watt)
        for i, r in enumerate(self.veriler):
            if self._build_key(r) == target:
                idx = i
                break

        if idx == -1:
            messagebox.showwarning("Uyarı", "Ürün bulunamadı.")
            return

        if not messagebox.askyesno("Onay", f"'{stok}' silinsin mi?"):
            return

        self._cleanup_trash()
        silinen = dict(self.veriler[idx])
        silinen["silinme_zamani"] = int(time.time())
        self.cop_kutusu.append(silinen)
        self.veriler.pop(idx)
        self._log(f"Ürün silindi: {stok}")

        self.urun_entry.delete(0, tk.END)
        self.stok_entry.delete(0, tk.END)
        self.cihaz_entry.delete(0, tk.END)
        self.fiyat_entry.delete(0, tk.END)

        self._save()
        self._refresh_models()
        self._refresh_table()
        if self._cop_refresh:
            self._cop_refresh()
        messagebox.showinfo("Başarılı", "Ürün çöp kutusuna atıldı. 1 saat sonra silinecek.")

    def _add_to_cart(self):
        secili = self.tablo.selection()
        if not secili:
            messagebox.showwarning("Uyarı", "Sepete eklemek için tablodan bir ürün seçin.")
            return

        if len(secili) == 1:
            vals = self.tablo.item(secili[0], "values")
            urun, stok, cihaz = vals[0], vals[1], vals[2]
            rec = self._find_record(urun, stok, cihaz)
            fiyat = record_get(rec, "Satis Fiyati") if rec else ""
            temiz_fiyat = parse_price(fiyat) if fiyat else "0"
            if not temiz_fiyat:
                temiz_fiyat = "0"

            miktar_pencere = tk.Toplevel(self.app)
            miktar_pencere.title("Miktar")
            miktar_pencere.geometry("300x150")
            miktar_pencere.resizable(False, False)
            miktar_pencere.transient(self.app)
            miktar_pencere.grab_set()
            miktar_pencere.update_idletasks()
            x = self.app.winfo_x() + (self.app.winfo_width() - 300) // 2
            y = self.app.winfo_y() + (self.app.winfo_height() - 150) // 2
            miktar_pencere.geometry(f"+{x}+{y}")

            tk.Label(miktar_pencere, text=f"{stok}", font=("Arial", 11, "bold")).pack(pady=(15, 5))
            tk.Label(miktar_pencere, text="Miktar:", font=("Arial", 10)).pack()

            miktar_var = tk.StringVar(value="1")
            miktar_entry = tk.Entry(miktar_pencere, textvariable=miktar_var, width=10, font=("Arial", 12))
            miktar_entry.pack(pady=5)
            miktar_entry.select_range(0, tk.END)
            miktar_entry.focus()

            def onayla():
                try:
                    miktar = int(miktar_var.get().strip())
                    if miktar <= 0:
                        raise ValueError
                except ValueError:
                    messagebox.showwarning("Uyarı", "Geçerli bir miktar girin.", parent=miktar_pencere)
                    return
                birim_fiyat = float(temiz_fiyat.replace(",", "."))
                toplam = round(birim_fiyat * miktar, 2)
                self.sepet.append({
                    "Kullanilan Urun": urun,
                    "Stok Adi": stok,
                    "Kullanilan Cihaz": cihaz,
                    "Birim Fiyat": birim_fiyat,
                    "Miktar": miktar,
                    "Toplam": toplam,
                })
                self._log(f"Sepete eklendi: {stok} x{miktar}")
                self._notify("Sepete Eklendi", f"{stok} x{miktar} - {toplam:,.2f} TL")
                self._save()
                miktar_pencere.destroy()
                self.tablo_alt.config(text=f"Toplam Kayıt: {len(self.veriler)}  |  Sepet: {len(self.sepet)} ürün")
                if self._sepet_refresh:
                    self._sepet_refresh()

            tk.Button(miktar_pencere, text="Ekle", font=("Arial", 10, "bold"), command=onayla).pack(pady=10)
            miktar_pencere.bind("<Return>", lambda e: onayla())
        else:
            eklenen = 0
            for item_id in secili:
                vals = self.tablo.item(item_id, "values")
                urun, stok, cihaz = vals[0], vals[1], vals[2]
                rec = self._find_record(urun, stok, cihaz)
                fiyat = record_get(rec, "Satis Fiyati") if rec else ""
                temiz_fiyat = parse_price(fiyat) if fiyat else "0"
                if not temiz_fiyat:
                    temiz_fiyat = "0"
                birim_fiyat = float(temiz_fiyat.replace(",", "."))
                self.sepet.append({
                    "Kullanilan Urun": urun,
                    "Stok Adi": stok,
                    "Kullanilan Cihaz": cihaz,
                    "Birim Fiyat": birim_fiyat,
                    "Miktar": 1,
                    "Toplam": birim_fiyat,
                })
                eklenen += 1
            self._log(f"Toplu sepete eklendi: {eklenen} ürün")
            self._notify("Sepete Eklendi", f"{eklenen} ürün sepete eklendi")
            self._save()
            self.tablo_alt.config(text=f"Toplam Kayıt: {len(self.veriler)}  |  Sepet: {len(self.sepet)} ürün")
            if self._sepet_refresh:
                self._sepet_refresh()

    def _open_cart(self):
        if self._sepet_pencere and self._sepet_pencere.winfo_exists():
            self._sepet_pencere.destroy()
            self._sepet_pencere = None
            self._sepet_refresh = None
            return
        self._log(f"Sepet açıldı ({len(self.sepet)} ürün)")

        pencere = tk.Toplevel(self.app)
        self._sepet_pencere = pencere
        pencere.title("Sepet")
        pencere.overrideredirect(True)
        pencere.attributes("-topmost", True)
        ax = self.app.winfo_x()
        ay = self.app.winfo_y()
        pencere.geometry(f"650x350+{ax}+{ay}")
        pencere.protocol("WM_DELETE_WINDOW", lambda: (setattr(self, '_sepet_pencere', None), setattr(self, '_sepet_refresh', None), pencere.destroy()))
        self._track_window(pencere)

        top_bar = tk.Frame(pencere, bg="#E8F0FE", cursor="hand2")
        top_bar.pack(side="top", fill="x")
        tk.Label(top_bar, text="  Sepet", font=("Arial", 12, "bold"), fg="#1F4E79", bg="#E8F0FE").pack(side="left", padx=5, pady=3)

        _drag_data = [0, 0]
        def _start_drag(event):
            _drag_data[0] = event.x_root - pencere.winfo_x()
            _drag_data[1] = event.y_root - pencere.winfo_y()
        def _do_drag(event):
            x = event.x_root - _drag_data[0]
            y = event.y_root - _drag_data[1]
            pencere.geometry(f"+{x}+{y}")
        top_bar.bind("<ButtonPress-1>", _start_drag)
        top_bar.bind("<B1-Motion>", _do_drag)
        self._sepet_count_label = tk.Label(top_bar, text=f"({len(self.sepet)} ürün)", font=("Arial", 10), fg="#666", bg="#E8F0FE")
        self._sepet_count_label.pack(side="left", padx=5)
        sepet_close = tk.Button(top_bar, text=" Kapat", image=self._btn_icons[8], compound="left",
                               font=("Segoe UI Emoji", 9), bd=0, cursor="hand2", bg="#E8F0FE", activebackground="#D0E4FA",
                               command=lambda: (pencere.destroy(), setattr(self, '_sepet_pencere', None), setattr(self, '_sepet_refresh', None)))
        sepet_close.pack(side="right", padx=5, pady=3)

        cols = ("urun", "stok", "cihaz", "birim", "miktar", "toplam")
        tablo = ttk.Treeview(pencere, columns=cols, show="headings", height=10, selectmode="browse")
        tablo.heading("urun", text="Ürün Adı")
        tablo.heading("stok", text="Stok Adı")
        tablo.heading("cihaz", text="Cihaz")
        tablo.heading("birim", text="Birim Fiyat")
        tablo.heading("miktar", text="Miktar")
        tablo.heading("toplam", text="Toplam")

        tablo.column("urun", width=180)
        tablo.column("stok", width=120)
        tablo.column("cihaz", width=80)
        tablo.column("birim", width=70, anchor="e")
        tablo.column("miktar", width=50, anchor="center")
        tablo.column("toplam", width=80, anchor="e")

        tablo.pack(fill="both", expand=True, padx=5, pady=(5, 0))

        alt = tk.Frame(pencere, bg="white")
        alt.pack(fill="x", padx=5, pady=5)

        toplam_label = tk.Label(alt, text="Toplam: 0 TL", font=("Arial", 11, "bold"), bg="white", fg="#1F4E79")
        toplam_label.pack(side="left")

        def refresh_cart():
            for item in tablo.get_children():
                tablo.delete(item)
            toplam = 0
            for s in self.sepet:
                tablo.insert("", "end", values=(
                    s["Kullanilan Urun"],
                    s["Stok Adi"],
                    s.get("Kullanilan Cihaz", ""),
                    f'{s["Birim Fiyat"]:.2f}'.replace(".", ","),
                    s["Miktar"],
                    f'{s["Toplam"]:.2f}'.replace(".", ","),
                ))
                toplam += s["Toplam"]
            toplam_label.config(text=f"Toplam: {toplam:,.2f} TL".replace(",", " ").replace(".", ",").replace(" ", "."))
            if hasattr(self, '_sepet_count_label') and self._sepet_count_label:
                self._sepet_count_label.config(text=f"({len(self.sepet)} ürün)")
            self.tablo_alt.config(text=f"Toplam Kayıt: {len(self.veriler)}  |  Sepet: {len(self.sepet)} ürün")

        self._sepet_refresh = refresh_cart
        refresh_cart()

        def sepetten_cikar():
            secili = tablo.selection()
            if not secili:
                messagebox.showwarning("Uyarı", "Çıkarmak için bir ürün seçin.", parent=pencere)
                return
            children = tablo.get_children()
            try:
                idx = list(children).index(secili[0])
            except ValueError:
                return
            if 0 <= idx < len(self.sepet):
                cikan = self.sepet.pop(idx)
                self._log(f"Sepetten çıkarıldı: {cikan['Stok Adi']}")
                self._save()
                refresh_cart()

        def sepeti_temizle():
            if not self.sepet:
                return
            if messagebox.askyesno("Onay", "Sepet temizlensin mi?", parent=pencere):
                self._log("Sepet temizlendi")
                self.sepet.clear()
                refresh_cart()

        btn_frame = tk.Frame(pencere, bg="white")
        btn_frame.pack(fill="x", padx=5, pady=(0, 5))

        btn_inner = tk.Frame(btn_frame, bg="white")
        btn_inner.pack(anchor="center")

        tk.Button(btn_inner, text=" Çıkar", image=self._btn_icons[7], compound="left", font=("Segoe UI Emoji", 9), command=sepetten_cikar).pack(side="left", padx=3)
        tk.Button(btn_inner, text=" Temizle", image=self._btn_icons[10], compound="left", font=("Segoe UI Emoji", 9), command=sepeti_temizle).pack(side="left", padx=3)
        def open_offer_and_close_cart():
            if not self.sepet:
                messagebox.showwarning("Uyarı", "Sepet boş.")
                return
            pencere.destroy()
            self._sepet_pencere = None
            self._sepet_refresh = None
            self._create_offer()
        tk.Button(btn_inner, text=" Teklif Oluştur", image=self._btn_icons[11], compound="left", font=("Segoe UI Emoji", 9), command=open_offer_and_close_cart).pack(side="left", padx=3)

    def _track_window(self, win):
        if not hasattr(self, '_tracked_windows'):
            self._tracked_windows = set()
            def _on_unmap(event):
                if event.widget == self.app:
                    try:
                        if self.app.state() != "iconic":
                            return
                    except Exception:
                        return
                    for w in list(self._tracked_windows):
                        try:
                            if w.winfo_exists():
                                w.destroy()
                        except Exception:
                            pass
                    self._tracked_windows.clear()
                    if hasattr(self, '_sepet_pencere'):
                        self._sepet_pencere = None
                        self._sepet_refresh = None
                    if hasattr(self, '_teklif_pencere'):
                        self._teklif_pencere = None
                    if hasattr(self, '_cop_pencere'):
                        self._cop_pencere = None
            self.app.bind("<Unmap>", _on_unmap)
            def _on_focus_out(event):
                try:
                    if self.app.state() == "iconic":
                        return
                except Exception:
                    return
                self._destroy_tracked()
            self.app.bind("<FocusOut>", _on_focus_out)
        self._tracked_windows.add(win)
        def _on_child_focus_out(event):
            self.app.after(50, self._destroy_tracked)
        win.bind("<FocusOut>", _on_child_focus_out)
        def _track():
            if not win.winfo_exists():
                self._tracked_windows.discard(win)
                return
            try:
                ax = self.app.winfo_x()
                ay = self.app.winfo_y()
                cx = win.winfo_x()
                cy = win.winfo_y()
                dx = ax - getattr(self, '_last_app_x', ax)
                dy = ay - getattr(self, '_last_app_y', ay)
                if dx != 0 or dy != 0:
                    win.geometry(f"+{cx + dx}+{cy + dy}")
                self._last_app_x = ax
                self._last_app_y = ay
            except Exception:
                pass
            self.app.after(50, _track)
        self._last_app_x = self.app.winfo_x()
        self._last_app_y = self.app.winfo_y()
        _track()

    def _destroy_tracked(self):
        if not hasattr(self, '_tracked_windows'):
            return
        try:
            if self.app.state() == "iconic":
                return
        except Exception:
            return
        root = self.app.winfo_toplevel()
        try:
            focused = root.focus_get()
        except Exception:
            focused = None
        if focused is not None:
            return
        for w in list(self._tracked_windows):
            try:
                if w.winfo_exists():
                    w.destroy()
            except Exception:
                pass
        self._tracked_windows.clear()
        if hasattr(self, '_sepet_pencere'):
            self._sepet_pencere = None
            self._sepet_refresh = None
        if hasattr(self, '_teklif_pencere'):
            self._teklif_pencere = None
        if hasattr(self, '_cop_pencere'):
            self._cop_pencere = None

    def _create_offer(self):
        if not self.sepet:
            messagebox.showwarning("Uyarı", "Sepet boş.")
            return

        try:
            sepet_kopya = copy.deepcopy(self.sepet)
        except Exception:
            sepet_kopya = list(self.sepet)

        win = tk.Toplevel(self.app)
        win.title("Teklif Formu")
        win.resizable(False, False)

        top = tk.Frame(win)
        top.pack(fill="x", padx=20, pady=(10,0))
        tk.Label(top, text="Teklif Bilgileri", font=("Arial", 14, "bold"), fg="#1F4E79").pack()

        form = tk.Frame(win)
        form.pack(fill="x", padx=20, pady=10)

        tk.Label(form, text="Müşteri Adı:", font=("Arial", 10)).grid(row=0, column=0, sticky="e", pady=4, padx=(0,5))
        musteri_ad = tk.Entry(form, width=30, font=("Arial", 10))
        musteri_ad.grid(row=0, column=1, pady=4)

        tk.Label(form, text="Firma:", font=("Arial", 10)).grid(row=1, column=0, sticky="e", pady=4, padx=(0,5))
        musteri_firma = tk.Entry(form, width=30, font=("Arial", 10))
        musteri_firma.grid(row=1, column=1, pady=4)

        tk.Label(form, text="Telefon:", font=("Arial", 10)).grid(row=2, column=0, sticky="e", pady=4, padx=(0,5))
        musteri_tel = tk.Entry(form, width=30, font=("Arial", 10))
        musteri_tel.grid(row=2, column=1, pady=4)

        tk.Label(form, text="Adres:", font=("Arial", 10)).grid(row=3, column=0, sticky="e", pady=4, padx=(0,5))
        musteri_adres = tk.Entry(form, width=30, font=("Arial", 10))
        musteri_adres.grid(row=3, column=1, pady=4)

        tk.Label(form, text="Ödeme:", font=("Arial", 10)).grid(row=4, column=0, sticky="e", pady=4, padx=(0,5))
        odeme_kosul = ttk.Combobox(form, values=["Peşin", "Kredi Kartı", "EFT", "Havale"], state="readonly", width=28, font=("Arial", 10))
        odeme_kosul.grid(row=4, column=1, pady=4)
        odeme_kosul.set("Peşin")

        sep = tk.Frame(win, height=2, bg="#1F4E79")
        sep.pack(fill="x", padx=20, pady=(5,5))

        btn = tk.Button(win, text="TEKLIF OLUSTUR", command=None, font=("Arial", 13, "bold"), bg="#1F4E79", fg="white", activebackground="#2D6CA2", activeforeground="white", relief="flat", cursor="hand2", padx=30, pady=8)

        def kaydet():
            ad = musteri_ad.get().strip() or "Müşteri"
            firma = musteri_firma.get().strip()
            tel = musteri_tel.get().strip()
            adres = musteri_adres.get().strip()
            kosul = odeme_kosul.get().strip()

            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Teklif"

                DARK = "1F4E79"
                WHITE = "FFFFFF"
                LIGHT = "E8F0FE"
                GRAY = "F5F5F5"
                SUB = "666666"

                title_font = Font(name="Calibri", size=20, bold=True, color=DARK)
                sub_font = Font(name="Calibri", size=11, color=SUB)
                bold = Font(name="Calibri", size=10, bold=True, color="000000")
                normal = Font(name="Calibri", size=10, color="000000")
                small = Font(name="Calibri", size=9, color=SUB)
                big_bold = Font(name="Calibri", size=16, bold=True, color=DARK)
                big_normal = Font(name="Calibri", size=16, color="000000")
                total_font = Font(name="Calibri", size=13, bold=True, color=DARK)
                th_font = Font(name="Calibri", size=10, bold=True, color=WHITE)

                thin = Side(style="thin", color="CCCCCC")
                med = Side(style="medium", color=DARK)
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                hdr_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
                alt_fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
                total_fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")
                th_fill = PatternFill(start_color="2D6CA2", end_color="2D6CA2", fill_type="solid")

                center = Alignment(horizontal="center", vertical="center")
                left = Alignment(horizontal="left", vertical="center")
                right = Alignment(horizontal="right", vertical="center")
                wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

                tarih = time.strftime("%d.%m.%Y")
                try:
                    with open(TEKLIF_COUNTER, "r") as f:
                        sayac = int(f.read().strip()) + 1
                except (FileNotFoundError, ValueError):
                    sayac = 1
                with open(TEKLIF_COUNTER, "w") as f:
                    f.write(str(sayac))
                teklif_no = f"KNKLF{sayac:012d}"

                logo_path = os.path.join(APP_DIR, "logo.ico")
                if not os.path.exists(logo_path) and getattr(sys, 'frozen', False):
                    logo_path = os.path.join(BUNDLE_DIR, "logo.ico")
                if os.path.exists(logo_path):
                    try:
                        from openpyxl.drawing.image import Image as XlImage
                        img = XlImage(logo_path)
                        img.width = 80
                        img.height = 80
                        ws.add_image(img, "A1")
                    except Exception:
                        pass

                ws.merge_cells("B1:G1")
                ws["B1"] = "KAEN ELEKTRİK ELEKTRONİK"
                ws["B1"].font = Font(name="Calibri", size=20, bold=True, color=DARK)
                ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 35

                ws.merge_cells("B2:G2")
                ws["B2"] = "ONVO SCOOTER YETKİLİ SERVİSİ"
                ws["B2"].font = Font(name="Calibri", size=12, bold=True, color=DARK)
                ws["B2"].alignment = center

                row = 4
                ws.merge_cells(f"A{row}:G{row}")
                ws.row_dimensions[row].height = 5
                for c in range(1, 8):
                    ws.cell(row=row, column=c).fill = hdr_fill

                row = 5
                ws.merge_cells(f"A{row}:C{row}")
                ws[f"A{row}"] = f"Teklif No: {teklif_no}"
                ws[f"A{row}"].font = bold
                ws.merge_cells(f"E{row}:G{row}")
                ws[f"E{row}"] = f"Tarih: {tarih}"
                ws[f"E{row}"].font = Font(name="Calibri", size=16, bold=True, color=DARK)
                ws[f"E{row}"].alignment = right

                row = 7
                bilgiler = [
                    ("Ad Soyad:", ad),
                    ("Firma:", firma),
                    ("Telefon:", tel),
                    ("Adres:", adres),
                    ("Ödeme:", kosul),
                ]
                for etiket, deger in bilgiler:
                    ws.merge_cells(f"A{row}:B{row}")
                    ws[f"A{row}"] = etiket
                    ws[f"A{row}"].font = big_bold
                    ws[f"A{row}"].alignment = left
                    ws.merge_cells(f"C{row}:G{row}")
                    ws[f"C{row}"] = deger
                    ws[f"C{row}"].font = big_normal
                    ws[f"C{row}"].alignment = left
                    row += 1

                row += 1
                ws.merge_cells(f"A{row}:G{row}")
                for c in range(1, 8):
                    ws.cell(row=row, column=c).fill = hdr_fill

                row += 1
                headers = ["#", "Ürün Adı", "Stok Adı", "Cihaz", "Birim Fiyat (TL)", "Miktar", "Toplam (TL)"]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=h)
                    cell.font = th_font
                    cell.fill = th_fill
                    cell.alignment = center
                    cell.border = border

                row += 1
                toplam = 0
                for i, s in enumerate(sepet_kopya, 1):
                    fill = alt_fill if i % 2 == 0 else PatternFill()
                    ws.cell(row=row, column=1, value=i).font = normal
                    ws.cell(row=row, column=2, value=s["Kullanilan Urun"]).font = normal
                    ws.cell(row=row, column=3, value=s["Stok Adi"]).font = normal
                    ws.cell(row=row, column=4, value=s.get("Kullanilan Cihaz", "")).font = normal
                    bf = s["Birim Fiyat"]
                    ws.cell(row=row, column=5, value=f'{bf:,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")).font = normal
                    ws.cell(row=row, column=6, value=s["Miktar"]).font = normal
                    tp = s["Toplam"]
                    ws.cell(row=row, column=7, value=f'{tp:,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")).font = normal
                    for c in range(1, 8):
                        cell = ws.cell(row=row, column=c)
                        cell.border = border
                        cell.fill = fill
                        cell.alignment = center
                    toplam += tp
                    row += 1

                row += 1
                ws.merge_cells(f"A{row}:F{row}")
                ws[f"A{row}"] = "GENEL TOPLAM"
                ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True, color=WHITE)
                ws[f"A{row}"].alignment = right
                ws[f"A{row}"].fill = hdr_fill
                ws.cell(row=row, column=7, value=f'{toplam:,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".") + " TL").font = Font(name="Calibri", size=13, bold=True, color=WHITE)
                ws.cell(row=row, column=7).alignment = center
                ws.cell(row=row, column=7).fill = hdr_fill
                for c in range(1, 7):
                    ws.cell(row=row, column=c).fill = hdr_fill
                ws.row_dimensions[row].height = 28

                row += 2
                ws.merge_cells(f"A{row}:G{row}")
                ws[f"A{row}"] = "KOSULLAR"
                ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color=WHITE)
                ws[f"A{row}"].fill = hdr_fill
                ws[f"A{row}"].alignment = left
                for c in range(1, 8):
                    ws.cell(row=row, column=c).fill = hdr_fill

                kosullar = [
                    f"1. Bu teklif {tarih} tarihinden itibaren 15 gün geçerlidir.",
                    f"2. Ödeme: {kosul}" if kosul else "",
                    "3. Fiyatlara KDV dahildir.",
                    "4. Stok durumuna göre teslimat süresi değişiklik gösterebilir.",
                ]
                for k in kosullar:
                    if not k:
                        continue
                    row += 1
                    ws.merge_cells(f"A{row}:G{row}")
                    ws[f"A{row}"] = k
                    ws[f"A{row}"].font = small
                    ws[f"A{row}"].alignment = wrap

                row += 2
                ws.merge_cells(f"A{row}:C{row}")
                ws[f"A{row}"] = "Müşteri İmzası: ___________________"
                ws[f"A{row}"].font = normal
                ws.merge_cells(f"E{row}:G{row}")
                ws[f"E{row}"] = "Kaen İmzası: ___________________"
                ws[f"E{row}"].font = normal
                ws[f"E{row}"].alignment = right

                ws.column_dimensions["A"].width = 6
                ws.column_dimensions["B"].width = 35
                ws.column_dimensions["C"].width = 38
                ws.column_dimensions["D"].width = 14
                ws.column_dimensions["E"].width = 14
                ws.column_dimensions["F"].width = 10
                ws.column_dimensions["G"].width = 16

                dosya_adi = f"Teklif_{ad}_{tarih.replace('.', '-')}.xlsx"
                kayit_yolu = os.path.join(os.path.expanduser("~"), "Desktop", dosya_adi)
                os.makedirs(TEKLIF_DIR, exist_ok=True)
                yedek_yolu = os.path.join(TEKLIF_DIR, dosya_adi)
                wb.save(kayit_yolu)
                wb.save(yedek_yolu)
                self._log(f"Teklif olusturuldu: {dosya_adi}")
                self._notify("Teklif Kaydedildi", f"Masaustune kaydedildi: {dosya_adi}")
                self.sepet.clear()
                
                # Save offer record
                teklif_kaydi = {
                    "no": teklif_no,
                    "tarih": tarih,
                    "musteri": ad,
                    "firma": firma,
                    "telefon": tel,
                    "adres": adres,
                    "odeme": kosul,
                    "toplam": toplam,
                    "urunler": [{"urun": s["Kullanilan Urun"], "stok": s["Stok Adi"], "cihaz": s.get("Kullanilan Cihaz", ""), "fiyat": s["Birim Fiyat"], "miktar": s["Miktar"], "toplam": s["Toplam"]} for s in sepet_kopya],
                    "dosya": dosya_adi,
                }
                teklifler = []
                if os.path.exists(TEKLIF_DATA):
                    try:
                        with open(TEKLIF_DATA, "r", encoding="utf-8") as f:
                            teklifler = json.load(f)
                    except Exception:
                        pass
                teklifler.insert(0, teklif_kaydi)
                with open(TEKLIF_DATA, "w", encoding="utf-8") as f:
                    json.dump(teklifler, f, ensure_ascii=False, indent=2)
                threading.Thread(target=lambda: requests.put(f"{FIREBASE_URL}/teklifler.json", json=teklifler, timeout=10), daemon=True).start()
                
                messagebox.showinfo("Basarili", f"Teklif kaydedildi:\n{kayit_yolu}")
                win.destroy()
            except Exception as e:
                import traceback
                messagebox.showerror("Hata", f"Teklif olusturulamadi:\n{traceback.format_exc()}")

        btn.configure(command=kaydet)
        btn.pack(pady=15)

        win.update_idletasks()
        sw = self.app.winfo_screenwidth()
        sh = self.app.winfo_screenheight()
        ww = win.winfo_width()
        wh = win.winfo_height()
        win.geometry(f"+{(sw - ww) // 2}+{(sh - wh) // 2}")

    def _open_teklifler(self):
        if hasattr(self, '_teklif_pencere') and self._teklif_pencere and self._teklif_pencere.winfo_exists():
            self._teklif_pencere.destroy()
            self._teklif_pencere = None
            return

        teklifler = []
        if os.path.exists(TEKLIF_DATA):
            try:
                with open(TEKLIF_DATA, "r", encoding="utf-8") as f:
                    teklifler = json.load(f)
            except Exception:
                pass

        if not teklifler:
            try:
                r = requests.get(f"{FIREBASE_URL}/teklifler.json", timeout=10)
                if r.status_code == 200:
                    remote_teklifler = r.json()
                    if isinstance(remote_teklifler, list) and remote_teklifler:
                        teklifler = remote_teklifler
                        os.makedirs(os.path.dirname(TEKLIF_DATA), exist_ok=True)
                        with open(TEKLIF_DATA, "w", encoding="utf-8") as f:
                            json.dump(teklifler, f, ensure_ascii=False, indent=2)
            except Exception:
                pass

        if not teklifler:
            messagebox.showinfo("Teklifler", "Henüz kayıtlı teklif yok.")
            return

        win = tk.Toplevel(self.app)
        self._teklif_pencere = win
        win.title("Kayıtlı Teklifler")
        win.overrideredirect(True)
        win.attributes("-topmost", True)
        ax = self.app.winfo_x()
        ay = self.app.winfo_y()
        win.geometry(f"700x350+{ax}+{ay}")
        win.protocol("WM_DELETE_WINDOW", lambda: (setattr(self, '_teklif_pencere', None), win.destroy()))
        self._track_window(win)

        top_bar = tk.Frame(win, bg="#E8F0FE", cursor="hand2")
        top_bar.pack(side="top", fill="x")
        tk.Label(top_bar, text="  Kayıtlı Teklifler", font=("Arial", 10, "bold"), fg="#1F4E79", bg="#E8F0FE").pack(side="left", padx=5, pady=3)
        tk.Button(top_bar, text=" Kapat", image=self._close_red_img, compound="left", font=("Segoe UI Emoji", 9), bd=0, cursor="hand2", bg="#E8F0FE", activebackground="#D0E4FA",
                  command=lambda: (setattr(self, '_teklif_pencere', None), win.destroy())).pack(side="right", padx=5, pady=3)

        _drag_data = [0, 0]
        def _start_drag(event):
            _drag_data[0] = event.x_root - win.winfo_x()
            _drag_data[1] = event.y_root - win.winfo_y()
        def _do_drag(event):
            x = event.x_root - _drag_data[0]
            y = event.y_root - _drag_data[1]
            win.geometry(f"+{x}+{y}")
        top_bar.bind("<ButtonPress-1>", _start_drag)
        top_bar.bind("<B1-Motion>", _do_drag)

        cols = ("Teklif No", "Tarih", "Müşteri", "Tutar", "Dosya")
        tbl = ttk.Treeview(win, columns=cols, show="headings", height=10)
        tbl.heading("Teklif No", text="Teklif No")
        tbl.heading("Tarih", text="Tarih")
        tbl.heading("Müşteri", text="Müşteri")
        tbl.heading("Tutar", text="Tutar (TL)")
        tbl.heading("Dosya", text="Dosya")

        tbl.column("Teklif No", width=150)
        tbl.column("Tarih", width=80)
        tbl.column("Müşteri", width=150)
        tbl.column("Tutar", width=100, anchor="e")
        tbl.column("Dosya", width=250)

        tbl.pack(fill="both", expand=True, padx=10, pady=5)

        tbl.bind("<Double-1>", lambda e: indir())

        for t in teklifler:
            tutar = f'{t.get("toplam", 0):,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")
            tbl.insert("", "end", values=(t.get("no", ""), t.get("tarih", ""), t.get("musteri", ""), tutar, t.get("dosya", "")))

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=10)

        def indir():
            sel = tbl.selection()
            if not sel:
                messagebox.showwarning("Uyarı", "Bir teklif seçin.", parent=win)
                return
            vals = tbl.item(sel[0], "values")
            dosya_adi = vals[4]
            masaustu = os.path.join(os.path.expanduser("~"), "Desktop", dosya_adi)
            yedek = os.path.join(TEKLIF_DIR, dosya_adi)
            if os.path.exists(masaustu):
                os.startfile(masaustu)
            elif os.path.exists(yedek):
                os.startfile(yedek)
            else:
                idx = tbl.index(sel[0])
                t = teklifler[idx]
                pdf_b64 = t.get("pdf_base64", "")
                if pdf_b64:
                    try:
                        import base64
                        os.makedirs(TEKLIF_DIR, exist_ok=True)
                        pdf_path = os.path.join(TEKLIF_DIR, dosya_adi.replace(".xlsx", ".pdf"))
                        with open(pdf_path, "wb") as f:
                            f.write(base64.b64decode(pdf_b64))
                        os.startfile(pdf_path)
                    except Exception:
                        messagebox.showinfo("Bilgi", "PDF indirilemedi.\nTeklif tekrar oluşturulmalı.", parent=win)
                else:
                    messagebox.showinfo("Bilgi", "Dosya bulunamadı.\nTeklif tekrar oluşturulmalı.", parent=win)
                    return
            self._teklif_pencere = None
            win.destroy()

        def refresh_table():
            tbl.delete(*tbl.get_children())
            for t in teklifler:
                tutar = f'{t.get("toplam", 0):,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")
                tbl.insert("", "end", values=(t.get("no", ""), t.get("tarih", ""), t.get("musteri", ""), tutar, t.get("dosya", "")))

        def save_json():
            with open(TEKLIF_DATA, "w", encoding="utf-8") as f:
                json.dump(teklifler, f, ensure_ascii=False, indent=2)
            threading.Thread(target=lambda: requests.put(f"{FIREBASE_URL}/teklifler.json", json=teklifler, timeout=10), daemon=True).start()

        def sil():
            sel = tbl.selection()
            if not sel:
                messagebox.showwarning("Uyarı", "Bir teklif seçin.", parent=win)
                return
            vals = tbl.item(sel[0], "values")
            if not messagebox.askyesno("Onay", f"'{vals[2]}' teklifi silinsin mi?", parent=win):
                return
            idx = tbl.index(sel[0])
            if 0 <= idx < len(teklifler):
                teklifler.pop(idx)
                save_json()
                refresh_table()

        def duzenle():
            sel = tbl.selection()
            if not sel:
                messagebox.showwarning("Uyarı", "Bir teklif seçin.", parent=win)
                return
            vals = tbl.item(sel[0], "values")
            idx = tbl.index(sel[0])
            if idx < 0 or idx >= len(teklifler):
                return

            dw = tk.Toplevel(win)
            dw.title("Teklif Düzenle")
            dw.geometry("400x350")
            dw.resizable(False, False)
            dw.attributes("-topmost", True)
            dw.update_idletasks()
            px = dw.winfo_screenwidth()
            py = dw.winfo_screenheight()
            dw.geometry(f"+{(px-400)//2}+{(py-350)//2}")
            self._tracked_windows.add(dw)
            dw.bind("<Destroy>", lambda e: self._tracked_windows.discard(dw), add="+")

            tk.Label(dw, text="Teklif No:", font=("Arial", 10)).grid(row=0, column=0, sticky="e", padx=10, pady=4)
            tk.Label(dw, text="Müşteri:", font=("Arial", 10)).grid(row=1, column=0, sticky="e", padx=10, pady=4)
            tk.Label(dw, text="Firma:", font=("Arial", 10)).grid(row=2, column=0, sticky="e", padx=10, pady=4)
            tk.Label(dw, text="Telefon:", font=("Arial", 10)).grid(row=3, column=0, sticky="e", padx=10, pady=4)
            tk.Label(dw, text="Adres:", font=("Arial", 10)).grid(row=4, column=0, sticky="e", padx=10, pady=4)
            tk.Label(dw, text="Ödeme:", font=("Arial", 10)).grid(row=5, column=0, sticky="e", padx=10, pady=4)
            tk.Label(dw, text="Tarih:", font=("Arial", 10)).grid(row=6, column=0, sticky="e", padx=10, pady=4)

            no_var = tk.StringVar(value=teklifler[idx].get("no", ""))
            musteri_var = tk.StringVar(value=teklifler[idx].get("musteri", ""))
            firma_var = tk.StringVar(value=teklifler[idx].get("firma", ""))
            tel_var = tk.StringVar(value=teklifler[idx].get("telefon", ""))
            adres_var = tk.StringVar(value=teklifler[idx].get("adres", ""))
            odeme_var = tk.StringVar(value=teklifler[idx].get("odeme", ""))
            tarih_var = tk.StringVar(value=teklifler[idx].get("tarih", ""))

            tk.Entry(dw, textvariable=no_var, width=30).grid(row=0, column=1, pady=4)
            tk.Entry(dw, textvariable=musteri_var, width=30).grid(row=1, column=1, pady=4)
            tk.Entry(dw, textvariable=firma_var, width=30).grid(row=2, column=1, pady=4)
            tk.Entry(dw, textvariable=tel_var, width=30).grid(row=3, column=1, pady=4)
            tk.Entry(dw, textvariable=adres_var, width=30).grid(row=4, column=1, pady=4)
            odeme_combo = ttk.Combobox(dw, textvariable=odeme_var, values=["Peşin", "Kredi Kartı", "EFT", "Havale"], state="readonly", width=28)
            odeme_combo.grid(row=5, column=1, pady=4)
            tk.Entry(dw, textvariable=tarih_var, width=30).grid(row=6, column=1, pady=4)

            def kaydet():
                old_no = teklifler[idx].get("no", "")
                old_musteri = teklifler[idx].get("musteri", "")
                old_tarih = teklifler[idx].get("tarih", "")
                new_no = no_var.get().strip()
                new_musteri = musteri_var.get().strip()
                new_tarih = tarih_var.get().strip()
                
                teklifler[idx]["no"] = new_no
                teklifler[idx]["musteri"] = new_musteri
                teklifler[idx]["firma"] = firma_var.get().strip()
                teklifler[idx]["telefon"] = tel_var.get().strip()
                teklifler[idx]["adres"] = adres_var.get().strip()
                teklifler[idx]["odeme"] = odeme_var.get().strip()
                teklifler[idx]["tarih"] = new_tarih
                save_json()
                
                # Update Excel files
                dosya = teklifler[idx].get("dosya", "")
                new_firma = firma_var.get().strip()
                new_tel = tel_var.get().strip()
                new_adres = adres_var.get().strip()
                new_odeme = odeme_var.get().strip()
                updated_excel = False
                for base in [os.path.join(os.path.expanduser("~"), "Desktop"), TEKLIF_DIR]:
                    xl_path = os.path.join(base, dosya)
                    if os.path.exists(xl_path):
                        try:
                            wb = load_workbook(xl_path)
                            ws = wb.active
                            for row_cells in ws.iter_rows():
                                for cell in row_cells:
                                    if cell.value is None:
                                        continue
                                    val = str(cell.value)
                                    if "Teklif No:" in val and old_no:
                                        cell.value = val.replace(old_no, new_no)
                                    elif "Tarih:" in val and old_tarih:
                                        cell.value = val.replace(old_tarih, new_tarih)
                                    elif "Ad Soyad:" not in val and "Firma:" not in val and "Telefon:" not in val and "Adres:" not in val and "Ödeme:" not in val:
                                        if old_musteri and old_musteri in val:
                                            cell.value = val.replace(old_musteri, new_musteri)
                            bilgiler_satir = None
                            for row_cells in ws.iter_rows():
                                for cell in row_cells:
                                    if cell.value and "Ad Soyad:" in str(cell.value):
                                        bilgiler_satir = cell.row
                                        break
                                if bilgiler_satir:
                                    break
                            if bilgiler_satir:
                                ws.cell(row=bilgiler_satir, column=3).value = new_musteri
                                ws.cell(row=bilgiler_satir + 1, column=3).value = new_firma
                                ws.cell(row=bilgiler_satir + 2, column=3).value = new_tel
                                ws.cell(row=bilgiler_satir + 3, column=3).value = new_adres
                                ws.cell(row=bilgiler_satir + 4, column=3).value = new_odeme
                            wb.save(xl_path)
                            wb.close()
                            updated_excel = True
                        except Exception as e:
                            messagebox.showwarning("Uyarı", f"Excel güncellenemedi:\n{xl_path}\n{str(e)[:100]}", parent=dw)
                        break
                if not updated_excel and dosya:
                    messagebox.showinfo("Bilgi", f"Excel dosyası bulunamadı:\n{dosya}\n\nDeğişiklikler sadece listede kaydedildi.", parent=dw)
                
                refresh_table()
                dw.destroy()

            tk.Button(dw, text="Kaydet", font=("Arial", 10, "bold"), command=kaydet).grid(row=7, column=0, columnspan=2, pady=15)

        def pdf_yap():
            sel = tbl.selection()
            if not sel:
                messagebox.showwarning("Uyarı", "Bir teklif seçin.", parent=win)
                return
            idx = tbl.index(sel[0])
            if idx < 0 or idx >= len(teklifler):
                return
            t = teklifler[idx]
            dosya = t.get("dosya", "")
            
            os.makedirs(TEKLIF_DIR, exist_ok=True)
            pdf_adi = dosya.replace(".xlsx", ".pdf") if dosya else f"Teklif_{t.get('no','')}.pdf"
            pdf_path = os.path.join(TEKLIF_DIR, pdf_adi)
            
            try:
                from fpdf import FPDF
                pdf = FPDF()
                pdf.add_page()
                # Try to load Calibri, fall back to built-in font
                try:
                    pdf.add_font("TkFont", "", "C:/Windows/Fonts/calibri.ttf", uni=True)
                    pdf.add_font("TkFont", "B", "C:/Windows/Fonts/calibrib.ttf", uni=True)
                    font_name = "TkFont"
                except Exception:
                    pdf.add_font("TkFont", "", "C:/Windows/Fonts/arial.ttf", uni=True)
                    pdf.add_font("TkFont", "B", "C:/Windows/Fonts/arialbd.ttf", uni=True)
                    font_name = "TkFont"
                
                # Logo
                logo_path = os.path.join(APP_DIR, "logo.ico")
                if not os.path.exists(logo_path) and getattr(sys, 'frozen', False):
                    logo_path = os.path.join(BUNDLE_DIR, "logo.ico")
                if os.path.exists(logo_path):
                    try:
                        tmp_path = os.path.join(os.environ.get("TEMP", "."), "_logo_temp.png")
                        from PIL import Image as PILImg
                        img = PILImg.open(logo_path)
                        img.save(tmp_path, "PNG")
                        pdf.image(tmp_path, x=10, y=8, w=25, h=25)
                        os.remove(tmp_path)
                    except Exception:
                        pass

                pdf.set_font(font_name, "B", 18)
                pdf.cell(0, 10, "KAEN ELEKTRIK ELEKTRONIK", ln=True, align="C")
                pdf.set_font(font_name, "B", 12)
                pdf.cell(0, 8, "ONVO SCOOTER YETKILI SERVISI", ln=True, align="C")
                pdf.ln(4)
                
                pdf.set_font(font_name, "B", 12)
                pdf.cell(90, 7, f"Teklif No: {t.get('no','')}", ln=0)
                pdf.cell(0, 7, f"Tarih: {t.get('tarih','')}", ln=1, align="R")
                pdf.ln(2)
                
                info = [("Ad Soyad:", t.get('musteri','')), ("Firma:", t.get('firma','')),
                        ("Telefon:", t.get('telefon','')), ("Adres:", t.get('adres','')),
                        ("Odeme:", t.get('odeme',''))]
                for etiket, deger in info:
                    if deger:
                        pdf.set_font(font_name, "B", 11)
                        pdf.cell(25, 7, etiket)
                        pdf.set_font(font_name, "", 11)
                        pdf.cell(0, 7, str(deger), ln=1)
                pdf.ln(4)
                
                # Table header
                cols = [("Stok Adi", 110), ("Miktar", 20), ("Toplam", 35)]
                pdf.set_fill_color(31, 78, 121)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font(font_name, "B", 9)
                for txt, w in cols:
                    pdf.cell(w, 7, txt, border=1, fill=True, align="C")
                pdf.ln()
                
                pdf.set_text_color(0, 0, 0)
                urunler = t.get("urunler", [])
                genel_toplam = 0
                for i, u in enumerate(urunler):
                    if i % 2 == 0:
                        pdf.set_fill_color(232, 240, 254)
                    else:
                        pdf.set_fill_color(255, 255, 255)
                    pdf.set_font(font_name, "", 9)
                    bf = u.get("fiyat", 0)
                    miktar = u.get("miktar", 0)
                    toplam = u.get("toplam", 0)
                    genel_toplam += toplam
                    
                    bf_txt = f'{bf:,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")
                    tp_txt = f'{toplam:,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")
                    
                    pdf.cell(110, 6, u.get("stok", ""), border=1, fill=True, align="C")
                    pdf.cell(20, 6, str(miktar), border=1, fill=True, align="C")
                    pdf.cell(35, 6, tp_txt, border=1, fill=True, align="C")
                    pdf.ln()
                
                # Total
                pdf.set_fill_color(31, 78, 121)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font(font_name, "B", 10)
                pdf.cell(130, 8, "GENEL TOPLAM", border=1, fill=True, align="R")
                genel_txt = f'{genel_toplam:,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")
                pdf.cell(35, 8, genel_txt + " TL", border=1, fill=True, align="C")
                pdf.ln(10)
                
                pdf.set_text_color(0, 0, 0)
                pdf.set_font(font_name, "", 8)
                pdf.cell(0, 5, "Fiyatlara KDV dahildir. Teklif 15 gun gecerlidir.", ln=1)
                
                pdf.output(pdf_path)
                try:
                    import base64
                    with open(pdf_path, "rb") as f:
                        pdf_b64 = base64.b64encode(f.read()).decode("utf-8")
                    teklifler[idx]["pdf_base64"] = pdf_b64
                    with open(TEKLIF_DATA, "w", encoding="utf-8") as f:
                        json.dump(teklifler, f, ensure_ascii=False, indent=2)
                    threading.Thread(target=lambda: requests.put(f"{FIREBASE_URL}/teklifler.json", json=teklifler, timeout=15), daemon=True).start()
                except Exception:
                    pass
                os.startfile(pdf_path)
                self._teklif_pencere = None
                win.destroy()
            except Exception as e:
                messagebox.showwarning("Hata", f"PDF olusturulamadi:\n{str(e)[:100]}", parent=win)

        tk.Button(btn_frame, text=" İndir / Aç", font=("Arial", 10, "bold"), command=indir).pack(side="left", padx=5)
        tk.Button(btn_frame, text=" PDF", font=("Arial", 10), command=pdf_yap).pack(side="left", padx=5)
        tk.Button(btn_frame, text=" Düzenle", font=("Arial", 10), command=duzenle).pack(side="left", padx=5)
        tk.Button(btn_frame, text=" Sil", font=("Arial", 10), command=sil).pack(side="left", padx=5)
        tk.Button(btn_frame, text=" Kapat", font=("Arial", 10), command=win.destroy).pack(side="left", padx=5)

    def _toggle_login(self):
        if self._logged_in:
            self._logged_in = False
            self._header_login.config(text=" Giriş ", fg="#1F4E79")
            self._btn_excel.grid_remove()
            self._btn_veri.grid_remove()
        else:
            from tkinter import simpledialog
            pw = simpledialog.askstring("Giriş", "Şifre:", parent=self.app)
            if pw and pw == self._admin_pw:
                self._logged_in = True
                self._header_login.config(text=" Çıkış ", fg="red")
                self._btn_excel.grid()
                self._btn_veri.grid()
            elif pw is not None:
                messagebox.showwarning("Hata", "Geçersiz şifre!")

    def _import_excel(self):
        path = filedialog.askopenfilename(
            title="Excel dosyası seç",
            initialdir=EXCEL_DIR,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        hedef = os.path.join(EXCEL_DIR, os.path.basename(path))
        if os.path.normpath(path) != os.path.normpath(hedef):
            shutil.copy2(path, hedef)
        source_excel = os.path.join(os.path.dirname(EXCEL_DIR), "excel")
        if os.path.normpath(source_excel) != os.path.normpath(EXCEL_DIR):
            os.makedirs(source_excel, exist_ok=True)
            src_hedef = os.path.join(source_excel, os.path.basename(path))
            if os.path.normpath(path) != os.path.normpath(src_hedef):
                shutil.copy2(path, src_hedef)
        self._log(f"Excel import: {os.path.basename(path)}")
        self._merge_workbook(hedef)

    def _merge_workbook(self, path):
        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Hata", f"Excel okunamadı:\n{e}")
            return

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messagebox.showwarning("Uyarı", "Dosyada veri yok.")
            return

        headers = ["" if c is None else str(c).strip() for c in rows[0]]
        imported = []
        for row in rows[1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            rec = build_record(row, headers)
            if any(rec.values()):
                imported.append(rec)

        if not imported:
            messagebox.showwarning("Uyarı", "Eklenebilir kayıt yok.")
            return

        existing_lookup = {self._build_key(r): r for r in self.veriler}
        trash_keys = {self._build_key(r) for r in self.cop_kutusu}
        for rec in imported:
            k = self._build_key(rec)
            if k in trash_keys:
                continue
            if k in existing_lookup:
                old = existing_lookup[k]
                fp = rec.get("Satis Fiyati", "")
                if fp:
                    record_set(old, "Satis Fiyati", self._markup(fp))
            else:
                rec["Satis Fiyati"] = self._markup(rec.get("Satis Fiyati", ""))
                self.veriler.append(rec)

        self._save()
        self._refresh_models()
        self._refresh_table()
        self._log(f"Excel içe aktarıldı: {len(imported)} kayıt - {os.path.basename(path)}")
        messagebox.showinfo("Başarılı", f"{len(imported)} kayıt aktarıldı.")

    def _reload_excel(self):
        if not messagebox.askyesno("Onay", "Excel klasöründeki tüm dosyalar işlenecek.\nTüm fiyatlar Excel'den güncellenecek.\nDevam edilsin mi?"):
            return
        old_count = len(self.veriler)
        self._read_excel_folder()
        new_count = len(self.veriler)
        self._refresh_models()
        self._refresh_table()
        self._log(f"Excel işlendi ({old_count} -> {new_count} kayıt)")
        messagebox.showinfo("Tamam", f"{old_count} kayıt vardı.\nŞimdi {new_count} kayıt var.\n{new_count - old_count} yeni kayıt eklendi.")

    def _read_excel_folder(self):
        if not os.path.isdir(EXCEL_DIR):
            self.durum.config(text="Excel klasörü bulunamadı.")
            return

        saved = {"veriler": self.veriler, "cop_kutusu": self.cop_kutusu}
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            saved = data if isinstance(data, dict) else saved
        except Exception:
            pass

        saved_records = saved.get("veriler", [])
        self.cop_kutusu = saved.get("cop_kutusu", [])
        saved_lookup = {self._build_key(r): r for r in saved_records}
        trash_keys = {self._build_key(r) for r in self.cop_kutusu}

        self.veriler = []
        seen = set()

        for dosya in sorted(os.listdir(EXCEL_DIR)):
            if not dosya.lower().endswith(".xlsx") or dosya.startswith("~$"):
                continue
            try:
                wb = load_workbook(os.path.join(EXCEL_DIR, dosya), data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = ["" if c is None else str(c).strip() for c in rows[0]]
                for row in rows[1:]:
                    if not row or all(c is None or str(c).strip() == "" for c in row):
                        continue
                    rec = build_record(row, headers)
                    if not any(rec.values()):
                        continue
                    k = self._build_key(rec)
                    if k in trash_keys or k in seen:
                        continue
                    seen.add(k)

                    excel_fp = rec.get("Satis Fiyati", "")
                    if excel_fp:
                        rec["Satis Fiyati"] = self._markup(excel_fp)
                    self.veriler.append(rec)
            except Exception:
                continue

        for r in saved_records:
            k = self._build_key(r)
            if k not in seen:
                seen.add(k)
                self.veriler.append(dict(r))

        self._save()
        self.durum.config(text=f"Toplam kayıt: {len(self.veriler)}")

    def _cleanup_trash(self):
        now = int(time.time())
        self.cop_kutusu = [e for e in self.cop_kutusu if isinstance(e, dict) and e.get("silinme_zamani", 0) + 3600 > now]

    def _periodic_cleanup(self):
        self._cleanup_trash()
        self._save()
        self._refresh_models()
        self._refresh_table()
        self.app.after(60000, self._periodic_cleanup)

    def _open_trash(self):
        if hasattr(self, '_cop_pencere') and self._cop_pencere and self._cop_pencere.winfo_exists():
            self._cop_pencere.destroy()
            self._cop_pencere = None
            return
        self._cleanup_trash()
        self._log(f"Çöp kutusu açıldı ({len(self.cop_kutusu)} ürün)")
        win = tk.Toplevel(self.app)
        self._cop_pencere = win
        win.title("Çöp Kutusu")
        win.overrideredirect(True)
        win.attributes("-topmost", True)
        ax = self.app.winfo_x()
        ay = self.app.winfo_y()
        win.geometry(f"650x300+{ax}+{ay}")
        win.lift()
        win.focus_force()
        self._track_window(win)

        top_bar = tk.Frame(win, bg="#E8F0FE", cursor="hand2")
        top_bar.pack(side="top", fill="x")
        tk.Label(top_bar, text="  Çöp Kutusu", font=("Arial", 10, "bold"), fg="#1F4E79", bg="#E8F0FE").pack(side="left", padx=5, pady=3)

        _drag_data = [0, 0]
        def _start_drag(event):
            win._user_dragging = True
            _drag_data[0] = event.x_root - win.winfo_x()
            _drag_data[1] = event.y_root - win.winfo_y()
        def _do_drag(event):
            x = event.x_root - _drag_data[0]
            y = event.y_root - _drag_data[1]
            win.geometry(f"+{x}+{y}")
        def _stop_drag(event):
            win._user_dragging = False
        top_bar.bind("<ButtonPress-1>", _start_drag)
        top_bar.bind("<B1-Motion>", _do_drag)
        top_bar.bind("<ButtonRelease-1>", _stop_drag)

        cols = ("Ürün", "Stok", "Model", "Fiyat")
        trash_tablo = ttk.Treeview(win, columns=cols, show="headings", height=5)
        for c in cols:
            trash_tablo.heading(c, text=c)
            trash_tablo.column(c, width=140, anchor="center")
        trash_tablo.pack(fill="both", expand=True, padx=10, pady=10)

        self._cop_refresh = None

        for idx, entry in enumerate(self.cop_kutusu):
            trash_tablo.insert("", "end", iid=str(idx), values=(
                entry.get("Kullanilan Urun", "") or entry.get("Kullanılan Ürün", ""),
                entry.get("Stok Adi", "") or entry.get("Stok Adı", ""),
                entry.get("Kullanilan Cihaz", "") or entry.get("Kullanılan Cihaz", ""),
                entry.get("Satis Fiyati", "") or entry.get("Satış Fiyatı", ""),
            ))

        def refresh_trash():
            for item in trash_tablo.get_children():
                trash_tablo.delete(item)
            for idx, entry in enumerate(self.cop_kutusu):
                trash_tablo.insert("", "end", iid=str(idx), values=(
                    entry.get("Kullanilan Urun", "") or entry.get("Kullanılan Ürün", ""),
                    entry.get("Stok Adi", "") or entry.get("Stok Adı", ""),
                    entry.get("Kullanilan Cihaz", "") or entry.get("Kullanılan Cihaz", ""),
                    entry.get("Satis Fiyati", "") or entry.get("Satış Fiyatı", ""),
                ))
        self._cop_refresh = refresh_trash

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=10)

        def restore():
            sel = trash_tablo.selection()
            if not sel:
                messagebox.showwarning("Uyarı", "Bir ürün seçin.")
                return
            idx = int(sel[0])
            if idx >= len(self.cop_kutusu):
                return
            item = self.cop_kutusu.pop(idx)
            item.pop("silinme_zamani", None)

            self.veriler.append(item)
            self._log(f"Geri alındı: {record_get(item, 'Stok Adi')}")
            self._save()
            self._refresh_models()
            self._refresh_table()
            self._select_by_values(record_get(item, "Kullanilan Urun"), record_get(item, "Stok Adi"), record_get(item, "Kullanilan Cihaz"))
            refresh_trash()
            messagebox.showinfo("Başarılı", "Ürün geri alındı.")

        def empty_trash():
            if not self.cop_kutusu:
                return
            if messagebox.askyesno("Onay", "Çöp kotasındaki tüm ürünler kalıcı olarak silinsin mi?", parent=win):
                count = len(self.cop_kutusu)
                self.cop_kutusu.clear()
                self._log(f"Çöp kutusu temizlendi ({count} ürün)")
                self._save()
                refresh_trash()
                messagebox.showinfo("Başarılı", "Çöp kutusu temizlendi.")

        def delete_selected():
            sel = trash_tablo.selection()
            if not sel:
                messagebox.showwarning("Uyarı", "Silmek için bir ürün seçin.")
                return
            idx = int(sel[0])
            if idx >= len(self.cop_kutusu):
                return
            silinen = self.cop_kutusu.pop(idx)
            self._log(f"Çöpten silindi: {record_get(silinen, 'Stok Adi')}")
            self._save()
            refresh_trash()

        tk.Button(btn_frame, text=" Geri Al", image=self._btn_icons[9], compound="left", font=("Segoe UI Emoji", 10), command=restore).pack(side="left", padx=5)
        tk.Button(btn_frame, text=" Seçili Sil", image=self._btn_icons[7], compound="left", font=("Segoe UI Emoji", 10), command=delete_selected).pack(side="left", padx=5)
        tk.Button(btn_frame, text=" Tümünü Sil", image=self._btn_icons[10], compound="left", font=("Segoe UI Emoji", 10), command=empty_trash).pack(side="left", padx=5)
        tk.Button(btn_frame, text=" Kapat", image=self._close_red_img, compound="left", font=("Segoe UI Emoji", 10), command=lambda: (setattr(self, '_cop_pencere', None), win.destroy())).pack(side="left", padx=5)


if __name__ == "__main__":
    app = StokUygulamasi()
    app.run()
